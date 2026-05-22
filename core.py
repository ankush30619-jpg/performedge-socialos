"""
Social Copy Studio — Core Engine
Handles: Brand loading, Gemini API calls, Excel I/O, Prompt building, Generation orchestration
"""
import json
import os
import re
import time
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
BRANDS_DIR = ROOT / "brands"
OUTPUT_DIR = ROOT / "output"
SETTINGS_FILE = ROOT / "settings.json"


# ─────────────────────────────────────────────────────────────────────────
# Settings
# ─────────────────────────────────────────────────────────────────────────
def load_settings() -> dict:
    if SETTINGS_FILE.exists():
        return json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
    return {
        "gemini_api_key": "", "model": "gemini-2.5-flash",
        "groq_api_key": "", "groq_model": "llama-3.3-70b-versatile",
        "apify_token": "", "news_api_key": "",
        "llm_provider": "groq",
    }


def save_settings(settings: dict) -> None:
    SETTINGS_FILE.write_text(json.dumps(settings, indent=2), encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────
# Brand loader
# ─────────────────────────────────────────────────────────────────────────
def list_brands() -> list[str]:
    return sorted([p.stem for p in BRANDS_DIR.glob("*.json")])


def load_brand(brand_key: str) -> dict:
    path = BRANDS_DIR / f"{brand_key}.json"
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    # Ensure scheduling schema exists
    if "last_pipeline_run" not in data:
        data["last_pipeline_run"] = None
    if "next_pipeline_due" not in data:
        data["next_pipeline_due"] = None
    return data


def save_brand(brand_key: str, data: dict) -> None:
    path = BRANDS_DIR / f"{brand_key}.json"
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────
# Prompt builder
# ─────────────────────────────────────────────────────────────────────────
def build_system_prompt(brand: dict) -> str:
    tone_rules = "\n".join(f"  • {r}" for r in brand.get("tone", {}).get("rules", []))
    do_not = "\n".join(f"  ✗ {r}" for r in brand.get("do_not_use", []))
    sig_phrases = ", ".join(f'"{p}"' for p in brand.get("signature_phrases", []))
    pain_points = "\n".join(f"  • {p}" for p in brand.get("audience", {}).get("pain_points", []))
    desires = "\n".join(f"  • {d}" for d in brand.get("audience", {}).get("desires", []))

    # Pull rich schema additions
    personas = brand.get("audience_personas", []) or []
    persona_block = ""
    if personas:
        lines = ["", "AUDIENCE PERSONAS (write copy TO these specific people, not 'everyone'):"]
        for p in personas[:3]:
            lines.append(f"  ▸ {p.get('name','')} ({p.get('age','')}) — {p.get('role','')[:120]}")
            if p.get("language_patterns"):
                lines.append(f"      They actually say: {p['language_patterns'][:200]}")
            if p.get("objections"):
                lines.append(f"      Their objections: {p['objections'][:200]}")
        persona_block = "\n".join(lines)

    cc = brand.get("cultural_context", {}) or {}
    cultural_block = ""
    if cc:
        lines = ["", "CULTURAL CONTEXT (bake this into every output):"]
        if cc.get("region"): lines.append(f"  Region: {cc['region']}")
        if cc.get("climate_reality"): lines.append(f"  Reality: {cc['climate_reality']}")
        if cc.get("dialect_phrases"):
            lines.append(f"  Use these dialect phrases naturally: " +
                        ", ".join(f'"{x}"' for x in cc["dialect_phrases"][:12]))
        if cc.get("local_references"):
            lines.append(f"  Reference these specifically when natural: " +
                        ", ".join(cc["local_references"][:8]))
        if cc.get("no_go_topics"):
            lines.append(f"  Cultural no-gos: {', '.join(cc['no_go_topics'][:5])}")
        cultural_block = "\n".join(lines)

    hooks = brand.get("hook_library", []) or []
    hook_block = ""
    if hooks:
        lines = ["", "BRAND HOOK LIBRARY (these are PROVEN openers — match this energy):"]
        for h in hooks[:10]:
            lines.append(f"  ▸ {h}")
        hook_block = "\n".join(lines)

    comp = brand.get("competitive_landscape", {}) or {}
    asp_block = ""
    if comp.get("aspirational_reference_brands"):
        lines = ["", "ASPIRATIONAL VOICE (channel THIS energy):"]
        for b in comp["aspirational_reference_brands"][:5]:
            lines.append(f"  ✓ {b}")
        if comp.get("anti_brands"):
            lines.append("AVOID sounding like:")
            for a in comp["anti_brands"][:4]:
                lines.append(f"  ✗ {a}")
        asp_block = "\n".join(lines)

    refs = brand.get("content_references", {}) or {}
    ref_block = ""
    if refs.get("examples_of_brand_voice_lines"):
        lines = ["", "EXAMPLES OF ON-BRAND VOICE (match this exact register):"]
        for x in refs["examples_of_brand_voice_lines"][:8]:
            lines.append(f"  ✓ \"{x}\"")
        ref_block = "\n".join(lines)
    if refs.get("format_losers_to_avoid"):
        ref_block += "\n\nDO NOT write in these styles:\n"
        for x in refs["format_losers_to_avoid"][:5]:
            ref_block += f"  ✗ {x}\n"

    # 6 universal hook patterns
    try:
        import format_templates as _ft
        universal_hooks_block = "\n\n" + _ft.hook_patterns_brief()
    except Exception:
        universal_hooks_block = ""

    products = brand.get("products_current") or brand.get("products_or_services") or brand.get("services") or []
    products_block = "\n".join(f"  ✓ {p}" for p in products)
    not_made = brand.get("products_NOT_made", [])
    not_made_block = "\n".join(f"  ✗ {p} (THIS BRAND DOES NOT MAKE/OFFER THIS)" for p in not_made) if not_made else ""
    product_facts = brand.get("product_facts_critical", [])
    product_facts_block = "\n".join(f"  ⚠ {f}" for f in product_facts) if product_facts else ""

    extra = ""
    if "compliance_notes" in brand and brand["compliance_notes"]:
        extra = "\n\nCOMPLIANCE (MUST FOLLOW):\n" + "\n".join(
            f"  ⚠ {c}" for c in brand["compliance_notes"]
        )

    product_guard = ""
    if products_block or not_made_block or product_facts_block:
        product_guard = f"""

═══════════════════════════════════════════════════════════════
PRODUCT FACTS — CRITICAL (do NOT hallucinate products)
THIS BRAND MAKES / OFFERS:
{products_block or '  (see brand profile)'}
{('THIS BRAND DOES NOT MAKE:' + chr(10) + not_made_block) if not_made_block else ''}
{('CRITICAL PRODUCT FACTS:' + chr(10) + product_facts_block) if product_facts_block else ''}
Never invent product names, features, or specifications. If the topic mentions a product, use ONLY the actual products listed above with their actual technology.
═══════════════════════════════════════════════════════════════"""

    return f"""You are the in-house social media copywriter for {brand['name']}. You are NOT a generic AI writer. You write copy that sounds unmistakably like this brand — scroll-stopping, culturally-specific, and built for reach.

You write for ONE specific persona at a time, in their actual language, opening with a pattern-interrupt that earns the next 5 seconds of attention.

═══════════════════════════════════════════════════════════════
BRAND: {brand['name']}
TAGLINE: {brand.get('tagline','')}
CATEGORY: {brand['category']}
LOCATION: {brand.get('location', 'N/A')}
═══════════════════════════════════════════════════════════════

AUDIENCE (overview)
{brand.get('audience', {}).get('primary', '')}

THEIR PAIN POINTS:
{pain_points}

WHAT THEY ACTUALLY WANT:
{desires}
{persona_block}
{cultural_block}

═══════════════════════════════════════════════════════════════
TONE OF VOICE
Personality: {brand.get('tone', {}).get('personality','')}
Voice: {brand.get('tone', {}).get('voice','')}

RULES YOU MUST FOLLOW:
{tone_rules}

SIGNATURE PHRASES (use naturally, never forced):
{sig_phrases}

NEVER DO:
{do_not}{extra}{product_guard}
{hook_block}
{asp_block}
{ref_block}
{universal_hooks_block}

═══════════════════════════════════════════════════════════════
STORY FORMULA: {brand.get('story_formula','Hook → Tension → Reveal → Proof → CTA')}

═══════════════════════════════════════════════════════════════
🔴 QUALITY BAR — auto-fail if ANY of these missed:
  1. Hook lands in FIRST 3 SECONDS — uses one of the 6 hook patterns above. No exceptions.
  2. Specific persona named (or addressed by their language patterns) — NOT "everyone"
  3. Cultural reference present when natural — dialect phrases, local moments, regional truth
  4. Mimics aspirational reference brands' energy — not corporate, not brochure
  5. ZERO banned phrases (elevate / transform / premium experience / etc.)
  6. Reel scripts = beat-by-beat with second timestamps. NEVER paragraphs.
  7. CTA uses psychology — urgency, opinion bait, FOMO, social proof — NEVER "DM us for more"

You will receive a post brief and return ONLY a JSON object — no preamble, no explanation."""


def _fetch_keywords_for_topic(topic: str, country: str = "in") -> list[str]:
    """Free keyword research — Google Autocomplete + variants. Returns up to 10."""
    try:
        import trend_scout
        suggestions = trend_scout.fetch_google_autocomplete(topic, country=country)
        return suggestions[:10]
    except Exception:
        return []


def _resolve_format_template(post: dict) -> tuple[str, str]:
    """Look up the viral format template for this post."""
    try:
        import format_templates as _ft
        return _ft.get_format_template(
            post.get("content_type", ""),
            post.get("notes", ""),
        )
    except Exception:
        return "", ""


def build_post_prompt(brand: dict, post: dict, competitor_playbook: str = "",
                       trends_brief: str = "",
                       use_keyword_research: bool = True) -> str:
    """Build the per-post user prompt."""
    fmt = (post.get("content_type") or "").strip().lower()
    is_ai_reel = ("ai reel" in fmt or "ai story" in fmt or "cgi" in fmt or
                  "ai video" in fmt or "story reel" in fmt or "ai-reel" in fmt)
    is_carousel = "carousel" in fmt and not is_ai_reel
    is_reel = ("reel" in fmt or "video" in fmt) and not is_ai_reel
    is_static = "static" in fmt or "graphic" in fmt or "post" in fmt
    is_story = ("story" in fmt or "stories" in fmt) and not is_ai_reel

    if is_ai_reel:
        format_instructions = """This is an AI / CGI STORY REEL. You are designing a 30–40 second emotional storytelling reel that will be generated using AI video tools (Runway Gen-3, Google Veo, Kling, Pika, Hailuo).

Build a *story-first* concept that:
  • Lands an emotional truth from this brand's audience (nostalgia, family, sacrifice, pride, joy, generations, small moments)
  • Uses the product as the catalyst — never the hero
  • Connects past ↔ present, or two human moments
  • Is cinematic and culturally resonant (think Cadbury / Surf Excel / Tanishq Indian ad-film calibre when relevant)
  • Could be made entirely with AI video clips stitched together

Output 5–7 SCENES. For each scene you MUST produce a directly-paste-into-Runway/Veo/Kling video prompt — describe the shot in vivid, cinematic detail (lighting, era, lens feel, mood, subject, action, wardrobe, location, color grade, aspect 9:16)."""
        slides_field = (
            '"concept_title": "Evocative 2-5 word title",\n  '
            '"emotional_core": "1-2 sentences capturing the emotional thesis of this reel",\n  '
            '"scenes": [\n    {\n      "scene_number": 1,\n      "duration": "0-5s",\n      '
            '"visual_description": "Plain-English description of what happens",\n      '
            '"voiceover": "Hinglish/English VO line for this beat (or \\\"[no VO]\\\")",\n      '
            '"on_screen_text": "Text on the screen (or \\\"[none]\\\")",\n      '
            '"ai_video_prompt": "FULL prompt ready to paste into Runway/Veo/Kling — cinematic, vivid, 9:16"\n    }\n  ],\n  '
            '"bgm_suggestion": "Specific mood description (e.g., \'Soft piano + tabla, nostalgic Indian-ad-film vibe\')"'
        )
    elif is_carousel:
        format_instructions = """This is a CAROUSEL post. Generate 5–7 slides of standalone copy. Each slide must:
- Have its own headline + 1-2 line body
- Build narrative momentum slide-to-slide
- Last slide = CTA slide

Return in slides[] array, each item: {"slide_number": 1, "headline": "...", "body": "..."}"""
        slides_field = '"slides": [{"slide_number": 1, "headline": "...", "body": "..."}, ...]'
    elif is_reel:
        format_instructions = """This is a REEL / VIDEO. Generate a full script broken into segments:
- HOOK (0–3 sec): scroll-stopper
- BODY (3–25 sec): story / content
- CLOSER (25–35 sec): payoff + CTA
Keep total under 38 seconds when read aloud. Include on-screen text suggestions in [brackets]."""
        slides_field = '"script": "HOOK (0-3s): ...\\n\\nBODY (3-25s): ...\\n\\nCLOSER (25-35s): ..."'
    elif is_story:
        format_instructions = """This is an INSTAGRAM STORY series. Generate 4–5 story slides as a sequence:
- Slide 1: Hook / question
- Middle slides: Quick info or poll/quiz prompt
- Final slide: CTA

Return in slides[] array."""
        slides_field = '"slides": [{"slide_number": 1, "headline": "...", "body": "..."}, ...]'
    else:
        format_instructions = """This is a STATIC GRAPHIC POST. Generate punchy on-creative copy:
- One bold headline (max 8 words)
- One supporting sub-headline (max 15 words)
- Optional: 2–3 micro-points if it's an infographic"""
        slides_field = '"creative_copy": {"headline": "...", "sub_headline": "...", "micro_points": ["...", "..."]}'

    occasion = post.get("festival") or ""
    occasion_line = f"\nOccasion / Festival: {occasion}" if occasion.strip() else ""
    audience_line = post.get("audience") or "primary brand audience"
    notes = post.get("notes") or ""
    notes_line = f"\nAdditional notes: {notes}" if notes.strip() else ""

    # ⭐ Resolve viral format template for this post
    format_name, format_template = _resolve_format_template(post)
    format_block = ""
    if format_template:
        format_block = (
            "\n\n═══ MANDATORY FORMAT STRUCTURE ═══\n"
            f"Format detected: {format_name}\n\n"
            f"{format_template}\n"
            "═══════════════════════════════════\n"
            "You MUST follow this structural template. The opening 0-3 sec, the beat structure, "
            "the voice notes — all of it. This is what separates 'AI default' from 'actually performs'."
        )

    competitor_block = ""
    if competitor_playbook.strip():
        competitor_block = (
            "\n\n═══ COMPETITOR INTELLIGENCE ═══\n"
            "Use this to DIFFERENTIATE — do NOT imitate. Identify what's saturated and avoid it. "
            "Lean into the underused angles. The goal is to stand out in this exact category right now.\n\n"
            f"{competitor_playbook}\n"
            "═══════════════════════════════"
        )

    trends_block = ""
    if trends_brief.strip():
        trends_block = (
            "\n\n═══ LIVE TRENDS SNAPSHOT ═══\n"
            "If a current trend, viral moment, or news cycle naturally fits this post — ride it. "
            "Reference specifics. Avoid generic seasonal mentions.\n\n"
            f"{trends_brief}\n"
            "═══════════════════════════════"
        )

    keyword_block = ""
    if use_keyword_research:
        kw_seeds = _fetch_keywords_for_topic(
            f"{post.get('main_topic', '')} {brand.get('name', '')}".strip()
        )
        if kw_seeds:
            keyword_block = (
                "\n\n═══ REAL KEYWORDS PEOPLE SEARCH (Google Autocomplete) ═══\n"
                "Weave these naturally into your caption and keywords list where appropriate. "
                "Do NOT just dump them — use them as signal for what audiences are actually asking.\n\n"
                + "\n".join(f"  • {k}" for k in kw_seeds) +
                "\n═════════════════════════════════════════════════"
            )

    return f"""Generate copy for this post:

Date: {post.get('date', 'TBD')}
Day: {post.get('day', '')}
Platform: {post.get('platform', 'Instagram, Facebook')}
Content Type: {post.get('content_type', '')}
Main Topic: {post.get('main_topic', '')}
Key Topic / Focus: {post.get('key_topic', '')}
Target Audience: {audience_line}{occasion_line}{notes_line}{format_block}{competitor_block}{trends_block}{keyword_block}

FORMAT INSTRUCTIONS:
{format_instructions}

OUTPUT — return ONLY this JSON object (no markdown, no preamble):

{{
  "hook_variants": [
    "First hook variant — scroll-stopper #1",
    "Second hook variant — different angle",
    "Third hook variant — different angle again"
  ],
  {slides_field},
  "caption": "Full Instagram/Facebook caption — story-driven, story-led, with line breaks for scannability. Should feel native to the platform. Include 1–2 emojis max where natural.",
  "keywords": ["seo keyword 1", "seo keyword 2", "...8 to 10 total"],
  "hashtags": ["#brand", "#topic", "#3-4-total-max"],
  "cta": "Specific call to action",
  "visual_note": "One-line direction for the designer about visuals — colors, layout cues, mood, shot style"
}}

Remember: write like ONLY {brand['name']} could have written this. Story first. Specifics over generic. Stop the scroll."""


# ─────────────────────────────────────────────────────────────────────────
# LLM Client factory (Gemini or Groq via llm_provider module)
# ─────────────────────────────────────────────────────────────────────────
import llm_provider as _llm

def make_llm_client(force_provider=None):
    """Build the active LLM client from current settings."""
    return _llm.make_client(load_settings(), force_provider=force_provider)


# Backwards compat alias
GeminiClient = _llm.GeminiClient


# ─────────────────────────────────────────────────────────────────────────
# Excel handlers
# ─────────────────────────────────────────────────────────────────────────
# Canonical input column mapping — case-insensitive, flexible
INPUT_COLUMN_ALIASES = {
    "date": ["date", "post date"],
    "day": ["day", "weekday"],
    "platform": ["platform", "platforms", "channel"],
    "content_type": ["content type", "format", "post type", "type"],
    "main_topic": ["main topic", "topic", "post topic"],
    "key_topic": ["key topic", "sub topic", "key focus", "focus", "key info (verified)"],
    "audience": ["audience", "target audience", "audience target"],
    "festival": ["festival", "occasion", "festival / occasion", "festival/occasion"],
    "notes": ["notes", "hook line", "remarks", "additional notes"],
}


def _normalize(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def read_input_calendar(path: str) -> list[dict]:
    """Read the input Excel and return a list of post dicts with normalized keys."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # Find header row (first row with multiple non-empty cells)
    header_row_idx = None
    for r in range(1, min(10, ws.max_row + 1)):
        cells = [c.value for c in ws[r]]
        non_empty = [c for c in cells if c not in (None, "")]
        if len(non_empty) >= 4:
            header_row_idx = r
            break
    if header_row_idx is None:
        raise ValueError("Couldn't find a header row in the input Excel.")

    headers = [_normalize(c.value) for c in ws[header_row_idx]]

    # Map columns
    col_map = {}
    for canonical, aliases in INPUT_COLUMN_ALIASES.items():
        for alias in aliases:
            for idx, h in enumerate(headers):
                if h == _normalize(alias):
                    col_map[canonical] = idx
                    break
            if canonical in col_map:
                break

    if "main_topic" not in col_map:
        raise ValueError(
            f"Could not find a 'Main Topic' column. Headers found: {headers}"
        )

    posts = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        if not any(row):
            continue
        post = {}
        for canonical, idx in col_map.items():
            val = row[idx] if idx < len(row) else None
            post[canonical] = str(val).strip() if val is not None else ""
        # skip if no main topic
        if not post.get("main_topic"):
            continue
        posts.append(post)

    return posts


def _format_slides(slides) -> str:
    if not slides:
        return ""
    lines = []
    for s in slides:
        n = s.get("slide_number", "?")
        h = s.get("headline", "")
        b = s.get("body", "")
        lines.append(f"SLIDE {n}: {h}\n{b}")
    return "\n\n".join(lines)


def _format_creative(creative) -> str:
    if not creative:
        return ""
    parts = []
    if creative.get("headline"):
        parts.append(f"HEADLINE: {creative['headline']}")
    if creative.get("sub_headline"):
        parts.append(f"SUB: {creative['sub_headline']}")
    if creative.get("micro_points"):
        parts.append("POINTS: " + " | ".join(creative["micro_points"]))
    return "\n".join(parts)


def _format_ai_reel(copy: dict) -> str:
    """Format an AI Reel concept output as a single readable cell."""
    parts = []
    if copy.get("concept_title"):
        parts.append(f"🎬 CONCEPT: {copy['concept_title']}")
    if copy.get("emotional_core"):
        parts.append(f"💡 EMOTIONAL CORE: {copy['emotional_core']}")
    if copy.get("scenes"):
        parts.append("")
        for s in copy["scenes"]:
            num = s.get("scene_number", "?")
            dur = s.get("duration", "")
            parts.append(f"━━━ SCENE {num}  ({dur}) ━━━")
            if s.get("visual_description"):
                parts.append(f"VISUAL: {s['visual_description']}")
            if s.get("voiceover"):
                parts.append(f"VOICEOVER: {s['voiceover']}")
            if s.get("on_screen_text"):
                parts.append(f"ON-SCREEN TEXT: {s['on_screen_text']}")
            if s.get("ai_video_prompt"):
                parts.append(f"🎞 AI VIDEO PROMPT (paste into Runway/Veo/Kling):")
                parts.append(f"   {s['ai_video_prompt']}")
            parts.append("")
    if copy.get("bgm_suggestion"):
        parts.append(f"🎵 BGM: {copy['bgm_suggestion']}")
    return "\n".join(parts)


def init_output_workbook(brand_name: str, posts: list[dict]):
    """Create the output workbook with input data filled + blank gen columns. Returns (wb, ws)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Generated Copy"

    title = f"{brand_name.upper()} — GENERATED SOCIAL CONTENT CALENDAR ({datetime.now().strftime('%d %b %Y')})"
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="003995")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "Date", "Day", "Platform", "Content Type", "Main Topic", "Key Topic",
        "Audience", "Festival / Occasion",
        "Hook Variant 1", "Hook Variant 2", "Hook Variant 3",
        "Script / Slides / Creative Copy", "Caption",
        "SEO Keywords", "Hashtags", "CTA", "Visual Note"
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0D1B3D")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 35

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for i, p in enumerate(posts, start=3):
        for col_idx, val in enumerate([
            p.get("date", ""), p.get("day", ""), p.get("platform", ""),
            p.get("content_type", ""), p.get("main_topic", ""), p.get("key_topic", ""),
            p.get("audience", ""), p.get("festival", "")
        ], start=1):
            c = ws.cell(row=i, column=col_idx, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor="F8EBE0")
        # Blank cells with formatting for the 9 generated columns
        for col_idx in range(9, 18):
            c = ws.cell(row=i, column=col_idx, value="")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor="F8EBE0")

    widths = [12, 10, 14, 14, 28, 24, 22, 16, 30, 30, 30, 60, 55, 30, 28, 22, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(3, len(posts) + 3):
        ws.row_dimensions[r].height = 220
    ws.freeze_panes = "A3"
    return wb, ws


def write_output_calendar(brand_name: str, posts_with_copy: list[dict], output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Generated Copy"

    # Title row
    title = f"{brand_name.upper()} — GENERATED SOCIAL CONTENT CALENDAR ({datetime.now().strftime('%d %b %Y')})"
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="003995")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "Date", "Day", "Platform", "Content Type", "Main Topic", "Key Topic",
        "Audience", "Festival / Occasion",
        "Hook Variant 1", "Hook Variant 2", "Hook Variant 3",
        "Script / Slides / Creative Copy",
        "Caption",
        "SEO Keywords", "Hashtags", "CTA", "Visual Note"
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0D1B3D")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 35

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for i, p in enumerate(posts_with_copy, start=3):
        copy = p.get("_copy", {})
        hooks = copy.get("hook_variants", ["", "", ""])
        hooks = (hooks + ["", "", ""])[:3]

        # Combine script/slides/creative/AI-reel into one column
        script_or_slides = ""
        if "scenes" in copy and copy["scenes"]:
            script_or_slides = _format_ai_reel(copy)
        elif "script" in copy and copy["script"]:
            script_or_slides = copy["script"]
        elif "slides" in copy and copy["slides"]:
            script_or_slides = _format_slides(copy["slides"])
        elif "creative_copy" in copy and copy["creative_copy"]:
            script_or_slides = _format_creative(copy["creative_copy"])

        keywords = ", ".join(copy.get("keywords", []) or [])
        hashtags = " ".join(copy.get("hashtags", []) or [])

        row_values = [
            p.get("date", ""),
            p.get("day", ""),
            p.get("platform", ""),
            p.get("content_type", ""),
            p.get("main_topic", ""),
            p.get("key_topic", ""),
            p.get("audience", ""),
            p.get("festival", ""),
            hooks[0],
            hooks[1],
            hooks[2],
            script_or_slides,
            copy.get("caption", ""),
            keywords,
            hashtags,
            copy.get("cta", ""),
            copy.get("visual_note", ""),
        ]
        for col_idx, val in enumerate(row_values, start=1):
            c = ws.cell(row=i, column=col_idx, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor="F8EBE0")

    # Column widths
    widths = [12, 10, 14, 14, 28, 24, 22, 16, 30, 30, 30, 60, 55, 30, 28, 22, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights — let them auto-grow but set a sane min
    for r in range(3, len(posts_with_copy) + 3):
        ws.row_dimensions[r].height = 220

    ws.freeze_panes = "A3"
    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────
# Generator (orchestrator)
# ─────────────────────────────────────────────────────────────────────────
def generate_calendar(
    brand_key: str,
    input_path: str,
    output_dir: str,
    progress_callback=None,
    log_callback=None,
    pause_event=None,
    stop_event=None,
) -> str:
    """
    Main entry point. Returns path to generated output file.
    progress_callback(current, total, post_label) called per post.
    log_callback(message) for status text.
    """
    settings = load_settings()
    brand = load_brand(brand_key)
    client = make_llm_client()
    system_prompt = build_system_prompt(brand)
    if log_callback:
        log_callback(f"Using LLM: {client.name.upper()} ({getattr(client, 'model_name', '?')})")

    # Load competitor playbook if it exists for this brand
    try:
        import competitor_analyzer
        competitor_playbook = competitor_analyzer.get_playbook_text(BRANDS_DIR, brand_key)
    except Exception:
        competitor_playbook = ""
    if competitor_playbook and log_callback:
        log_callback("✓ Competitor playbook loaded — will inject into prompts.")

    # Load trends snapshot if it exists
    try:
        import trend_scout
        trends_brief = trend_scout.get_trend_brief(BRANDS_DIR, brand_key)
    except Exception:
        trends_brief = ""
    if trends_brief and log_callback:
        log_callback("✓ Trends snapshot loaded — will inject into prompts.")

    if log_callback:
        log_callback(f"Reading calendar: {Path(input_path).name}")
    posts = read_input_calendar(input_path)
    if not posts:
        raise ValueError("No posts found in the input calendar.")
    if log_callback:
        log_callback(f"Found {len(posts)} posts. Starting generation…")

    # Set up output file UP FRONT so partial work is preserved
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^A-Za-z0-9_-]", "_", brand["name"])
    out_path = str(Path(output_dir) / f"{safe_name}_Generated_{timestamp}.xlsx")
    sidecar_path = out_path.replace(".xlsx", "_data.json")

    wb, ws = init_output_workbook(brand["name"], posts)
    wb.save(out_path)
    if log_callback:
        log_callback(f"Output file created: {Path(out_path).name}")

    was_paused_announced = False
    aborted = False
    consecutive_429 = 0
    MAX_CONSECUTIVE_429 = 3

    for idx, post in enumerate(posts, start=1):
        # ── Stop check ──
        if stop_event is not None and stop_event.is_set():
            if log_callback:
                log_callback(f"⏹ Stopped by user at post {idx}/{len(posts)}.")
            aborted = True
            break

        # ── Pause loop ──
        if pause_event is not None and pause_event.is_set():
            if log_callback:
                log_callback(f"⏸ Paused at post {idx}/{len(posts)}. Click Resume to continue, or Stop to abort.")
            was_paused_announced = True
            while pause_event.is_set():
                if stop_event is not None and stop_event.is_set():
                    break
                time.sleep(0.4)
            if stop_event is not None and stop_event.is_set():
                if log_callback:
                    log_callback("⏹ Stopped during pause.")
                aborted = True
                break
            if log_callback and was_paused_announced:
                log_callback("▶ Resumed.")
            was_paused_announced = False

        label = f"{post.get('date','')} — {post.get('main_topic','')[:50]}"
        if progress_callback:
            progress_callback(idx - 1, len(posts), label)
        if log_callback:
            log_callback(f"[{idx}/{len(posts)}] Generating: {label}")

        user_prompt = build_post_prompt(brand, post,
                                          competitor_playbook=competitor_playbook,
                                          trends_brief=trends_brief)
        try:
            copy = client.generate(system_prompt, user_prompt, log_callback=log_callback)
            consecutive_429 = 0
        except Exception as e:
            err_str = str(e)
            if log_callback:
                log_callback(f"  ⚠ Error on post {idx}: {e}")
            copy = {"caption": f"[GENERATION FAILED: {e}]"}
            if "429" in err_str or "quota" in err_str.lower() or "rate" in err_str.lower():
                consecutive_429 += 1
                if consecutive_429 >= MAX_CONSECUTIVE_429:
                    if log_callback:
                        log_callback(f"\n✗ {MAX_CONSECUTIVE_429} consecutive rate-limit / quota errors.")
                        log_callback("   Daily quota likely exhausted. Options:")
                        log_callback("   1) Settings → switch LLM Provider to Groq (or vice versa).")
                        log_callback("   2) Wait until quota resets (Gemini: 24h Pacific, Groq: rolls minute-by-minute).")
                        log_callback("   3) Click 'Resume Failed' later to finish remaining posts.")
                    post["_copy"] = copy
                    try:
                        _write_row_copy(ws, idx + 2, copy)
                        wb.save(out_path)
                    except Exception: pass
                    aborted = True
                    break
        post["_copy"] = copy

        # Incremental save — write this row to disk RIGHT NOW
        try:
            _write_row_copy(ws, idx + 2, copy)  # +2 because rows 1-2 are title/header
            wb.save(out_path)
        except Exception as e:
            if log_callback:
                log_callback(f"  ⚠ Failed to save row {idx}: {e}")

        if progress_callback:
            progress_callback(idx, len(posts), label)

        # Pacing — stay well under free-tier RPM/TPM limits between posts
        if idx < len(posts):
            # Pause-aware pacing — break early if user pauses or stops
            elapsed = 0.0
            while elapsed < 8.0:
                if stop_event is not None and stop_event.is_set():
                    break
                if pause_event is not None and pause_event.is_set():
                    break
                time.sleep(0.4)
                elapsed += 0.4

    # Sidecar JSON (always write, even on partial / aborted)
    try:
        with open(sidecar_path, "w", encoding="utf-8") as f:
            json.dump({"brand_key": brand_key, "posts": posts,
                       "aborted": aborted,
                       "completed_at": datetime.now().isoformat(timespec="seconds")},
                      f, indent=2, ensure_ascii=False, default=str)
    except Exception:
        pass

    if log_callback:
        if aborted:
            log_callback(f"⏹ Run ended early — partial output saved: {out_path}")
            log_callback("   Click 'Resume Failed' later to continue any rows that didn't generate.")
        else:
            log_callback(f"✓ Done! Output saved: {out_path}")
    return out_path


# ─────────────────────────────────────────────────────────────────────────
# Resume failed generations
# ─────────────────────────────────────────────────────────────────────────
# Output column index map (1-based)
_OUT_COL = {
    "date": 1, "day": 2, "platform": 3, "content_type": 4,
    "main_topic": 5, "key_topic": 6, "audience": 7, "festival": 8,
    "hook1": 9, "hook2": 10, "hook3": 11, "script_slides": 12,
    "caption": 13, "keywords": 14, "hashtags": 15, "cta": 16, "visual_note": 17,
}


def resume_failed_generations(output_xlsx_path: str,
                               progress_callback=None,
                               log_callback=None,
                               pause_event=None,
                               stop_event=None) -> tuple[int, int]:
    """
    Re-generate only the rows in an existing output Excel that have
    '[GENERATION FAILED' in the caption column. Writes updates back to the
    same Excel file. Returns (succeeded, still_failed).
    """
    settings = load_settings()
    wb = load_workbook(output_xlsx_path)
    ws = wb.active

    # Try to load sidecar JSON for full notes/context
    sidecar_path = output_xlsx_path.replace(".xlsx", "_data.json")
    sidecar = None
    if Path(sidecar_path).exists():
        try:
            sidecar = json.loads(Path(sidecar_path).read_text(encoding="utf-8"))
        except Exception:
            sidecar = None

    # Determine brand_key from sidecar or by best guess from filename
    brand_key = None
    if sidecar:
        brand_key = sidecar.get("brand_key")
    if not brand_key:
        # try matching by brand name in title (row 1)
        title = str(ws.cell(1, 1).value or "")
        for k in list_brands():
            try:
                b = load_brand(k)
                if b["name"].upper() in title.upper():
                    brand_key = k
                    break
            except Exception:
                continue
    if not brand_key:
        raise ValueError("Could not detect brand for this file. Make sure the original "
                         "sidecar JSON is present next to the Excel.")

    brand = load_brand(brand_key)
    client = make_llm_client()
    system_prompt = build_system_prompt(brand)
    if log_callback:
        log_callback(f"Using LLM: {client.name.upper()} ({getattr(client, 'model_name', '?')})")

    try:
        import competitor_analyzer
        competitor_playbook = competitor_analyzer.get_playbook_text(BRANDS_DIR, brand_key)
    except Exception:
        competitor_playbook = ""

    # Identify failed rows
    failed = []
    for r in range(3, ws.max_row + 1):
        caption = str(ws.cell(r, _OUT_COL["caption"]).value or "")
        if "GENERATION FAILED" in caption:
            # Build post dict from the row
            post = {
                "date": str(ws.cell(r, _OUT_COL["date"]).value or ""),
                "day": str(ws.cell(r, _OUT_COL["day"]).value or ""),
                "platform": str(ws.cell(r, _OUT_COL["platform"]).value or ""),
                "content_type": str(ws.cell(r, _OUT_COL["content_type"]).value or ""),
                "main_topic": str(ws.cell(r, _OUT_COL["main_topic"]).value or ""),
                "key_topic": str(ws.cell(r, _OUT_COL["key_topic"]).value or ""),
                "audience": str(ws.cell(r, _OUT_COL["audience"]).value or ""),
                "festival": str(ws.cell(r, _OUT_COL["festival"]).value or ""),
                "notes": "",
            }
            # Enrich with notes from sidecar if available (match by date + main_topic)
            if sidecar:
                for sp in sidecar.get("posts", []):
                    if (str(sp.get("date", "")).strip() == post["date"].strip() and
                            str(sp.get("main_topic", "")).strip() == post["main_topic"].strip()):
                        post["notes"] = sp.get("notes", "")
                        break
            failed.append((r, post))

    if not failed:
        if log_callback:
            log_callback("No failed rows found. Nothing to resume.")
        return (0, 0)

    if log_callback:
        log_callback(f"Found {len(failed)} failed row(s) to retry.")

    succeeded = 0
    still_failed = 0
    consecutive_429 = 0

    for idx, (row_num, post) in enumerate(failed, start=1):
        # Stop check
        if stop_event is not None and stop_event.is_set():
            if log_callback:
                log_callback(f"⏹ Stopped at row {idx}/{len(failed)}.")
            break
        # Pause loop
        if pause_event is not None and pause_event.is_set():
            if log_callback:
                log_callback(f"⏸ Paused at row {idx}/{len(failed)}. Click Resume.")
            while pause_event.is_set():
                if stop_event is not None and stop_event.is_set():
                    break
                time.sleep(0.4)
            if stop_event is not None and stop_event.is_set():
                break
            if log_callback:
                log_callback("▶ Resumed.")

        label = f"row {row_num}: {post['main_topic'][:48]}"
        if progress_callback:
            progress_callback(idx - 1, len(failed), label)
        if log_callback:
            log_callback(f"[{idx}/{len(failed)}] Retrying {label}")

        user_prompt = build_post_prompt(brand, post,
                                          competitor_playbook=competitor_playbook,
                                          trends_brief=trends_brief)
        try:
            copy = client.generate(system_prompt, user_prompt, log_callback=log_callback)
            _write_row_copy(ws, row_num, copy)
            wb.save(output_xlsx_path)  # save after each successful row
            succeeded += 1
            if log_callback:
                log_callback(f"     ✓ Updated row {row_num}")
        except Exception as e:
            still_failed += 1
            err_str = str(e)
            if log_callback:
                log_callback(f"     ⚠ Still failing row {row_num}: {e}")
            if "429" in err_str or "quota" in err_str.lower() or "rate" in err_str.lower():
                consecutive_429 += 1
                if consecutive_429 >= 3:
                    if log_callback:
                        log_callback("\n✗ 3 consecutive rate-limit / quota errors.")
                        log_callback("   Daily quota likely exhausted.")
                        log_callback("   Switch LLM Provider in Settings (Gemini ↔ Groq), or wait for reset.")
                        log_callback("   Click 'Resume Failed' again later.")
                    break

        if progress_callback:
            progress_callback(idx, len(failed), label)

        # Pause-aware pacing
        if idx < len(failed):
            elapsed = 0.0
            while elapsed < 8.0:
                if stop_event is not None and stop_event.is_set():
                    break
                if pause_event is not None and pause_event.is_set():
                    break
                time.sleep(0.4); elapsed += 0.4

    wb.save(output_xlsx_path)
    if log_callback:
        log_callback(f"\n✓ Resume complete — {succeeded} succeeded, {still_failed} still failing.")
    return (succeeded, still_failed)


def _write_row_copy(ws, row_num: int, copy: dict) -> None:
    """Write the generated copy fields into the given row of the output sheet."""
    hooks = copy.get("hook_variants", []) or []
    hooks = (list(hooks) + ["", "", ""])[:3]

    script_or_slides = ""
    if "scenes" in copy and copy["scenes"]:
        script_or_slides = _format_ai_reel(copy)
    elif "script" in copy and copy["script"]:
        script_or_slides = copy["script"]
    elif "slides" in copy and copy["slides"]:
        script_or_slides = _format_slides(copy["slides"])
    elif "creative_copy" in copy and copy["creative_copy"]:
        script_or_slides = _format_creative(copy["creative_copy"])

    keywords = ", ".join(copy.get("keywords", []) or [])
    hashtags = " ".join(copy.get("hashtags", []) or [])

    ws.cell(row=row_num, column=_OUT_COL["hook1"], value=hooks[0])
    ws.cell(row=row_num, column=_OUT_COL["hook2"], value=hooks[1])
    ws.cell(row=row_num, column=_OUT_COL["hook3"], value=hooks[2])
    ws.cell(row=row_num, column=_OUT_COL["script_slides"], value=script_or_slides)
    ws.cell(row=row_num, column=_OUT_COL["caption"], value=copy.get("caption", ""))
    ws.cell(row=row_num, column=_OUT_COL["keywords"], value=keywords)
    ws.cell(row=row_num, column=_OUT_COL["hashtags"], value=hashtags)
    ws.cell(row=row_num, column=_OUT_COL["cta"], value=copy.get("cta", ""))
    ws.cell(row=row_num, column=_OUT_COL["visual_note"], value=copy.get("visual_note", ""))

    # Re-apply formatting (alternating fill + wrap)
    for c in range(1, 18):
        cell = ws.cell(row=row_num, column=c)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
