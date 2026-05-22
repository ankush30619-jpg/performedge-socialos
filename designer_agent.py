"""
Designer Agent — generates premium social-media graphics for every non-video
post in a completed copy calendar.

Workflow per post:
  1. Read row from copy Excel (caption, visual_note, script/slides)
  2. Art Director (LLM): craft an ultra-detailed image prompt
     using brand identity + post content
  3. Freepik (Mystic / Imagen3): render the image
  4. Save to disk in IG-ready sizes

Carousels → one image per slide, consistent style.
Statics → 1 image (1080×1080).
Stories → 1 image (1080×1920).
Reels / AI Reels / Videos → SKIPPED.
"""
import json
import re
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import freepik_client
import core
import compositor


# ───────────────────────── Content type detection ─────────────────────────
AI_VIDEO_KEYS = {"ai reel", "ai story reel", "cgi reel", "ai video", "story reel", "ai-reel"}


def detect_content_format(content_type: str) -> str:
    """Returns: 'carousel', 'story', 'static', or 'video' (skip)."""
    ct = (content_type or "").lower().strip()
    if any(v in ct for v in AI_VIDEO_KEYS):
        return "video"
    if "reel" in ct or ct == "video":
        return "video"
    if "carousel" in ct:
        return "carousel"
    if "story" in ct or "stories" in ct:
        return "story"
    return "static"


# ───────────────────────── Art Director (LLM) ─────────────────────────
ART_DIRECTOR_SYSTEM = """You are an elite art director with 15 years designing premium social-media campaigns for major brands. Your craft is writing image-generation prompts that produce magazine-grade visuals — NOT generic stock-photo descriptions.

Every prompt you write must control:
  • SUBJECT — what's in frame, who's there, what's happening (specific, not generic)
  • COMPOSITION — rule of thirds, focal point, depth of field, leading lines
  • LIGHTING — direction (front/side/back), quality (hard/soft), color temperature (cool/warm/golden), mood (dramatic/serene/festive)
  • COLOR PALETTE — work the brand's exact hex codes into the prompt as adjectival color descriptors (e.g., "deep royal blue #003995 accents")
  • MOOD & EMOTION — register/atmosphere that fits the brand voice
  • STYLE — photographic / illustrative / editorial / cinematic / hyper-realistic / claymation — pick what fits THIS brand
  • DETAILS — wardrobe, props, location, era, weather, time of day, materials
  • TECHNICAL — depth, lens (e.g. 35mm, 85mm portrait), ISO feel, post-production look (color grade)
  • BRAND-WORLD CONSISTENCY — props, locations, people that feel ON BRAND

CRITICAL RULES:
  ✗ Do NOT mention text, typography, captions, words, or letters in the image (text is overlaid separately by designer)
  ✗ Do NOT describe the aspect ratio in the prompt — handled separately
  ✗ Avoid generic stock-photo language ("a man working on laptop, smiling")
  ✓ Use cinematic, evocative, art-direction language
  ✓ For carousel slides: keep style + palette + lens consistent across all slides in the same carousel
  ✗ DO NOT render the brand's logo, name, or any product PRODUCT-WITH-OVERLAY MODE.
  ✗ DO NOT render any brand logo, brand name, or readable text.
  ✗ When the brief says product-overlay mode: do NOT render the product itself. Instead, design a scene with intentional negative space in the FOREGROUND-CENTER or RIGHT-THIRD where a transparent product PNG will be composited. The scene must still feel complete — props, depth, lighting, mood — just leave the product hero zone uncluttered.
  ✗ Always design the TOP-LEFT corner with a clean, low-contrast area suitable for a logo overlay (sky, soft wall, gradient, blur — not faces or busy props).
  ✓ Use cinematic, evocative, art-direction language

Output ONLY this JSON shape:
{"image_prompt": "the full prompt, 80-180 words of dense visual direction"}"""


def craft_image_prompt(brand: dict, post: dict, slide_data: dict = None,
                       carousel_style_anchor: str = "",
                       product_overlay_mode: bool = False,
                       llm_client=None, log=None) -> tuple[str, str]:
    """
    Use the LLM (art director) to design the image prompt.
    Returns (prompt_text, style_anchor) where style_anchor is a short
    descriptor to keep carousel slides visually consistent.
    """
    slide_block = ""
    if slide_data:
        slide_block = f"""

CAROUSEL SLIDE CONTEXT
You're designing slide {slide_data['n']} of a multi-slide carousel.
This slide's headline: "{slide_data.get('headline', '')}"
This slide's body: "{slide_data.get('body', '')[:300]}"

{'STYLE ANCHOR FROM SLIDE 1 (must match for consistency): ' + carousel_style_anchor if carousel_style_anchor else 'You are designing slide 1 — set the visual style anchor that subsequent slides will inherit.'}"""

    products = (brand.get("products_current") or brand.get("products_or_services")
                or brand.get("services") or [])
    not_made = brand.get("products_NOT_made", [])

    brand_block = f"""BRAND: {brand.get('name', '')}
Category: {brand.get('category', '')}
Tagline: {brand.get('tagline', '')}
Visual identity: colors → primary {brand.get('colors',{}).get('primary','#000000')}, accent {brand.get('colors',{}).get('secondary','#FFFFFF')}, supporting {brand.get('colors',{}).get('accent','#666666')}
Personality: {brand.get('tone', {}).get('personality', '')}
Audience: {brand.get('audience', {}).get('primary', '')[:300]}
Products: {', '.join(products[:6])}
{('PRODUCT FACTS — DO NOT MISREPRESENT: ' + ' | '.join(brand.get('product_facts_critical', [])[:4])) if brand.get('product_facts_critical') else ''}
{'DOES NOT MAKE: ' + ', '.join(not_made) if not_made else ''}"""

    overlay_directive = ""
    if product_overlay_mode:
        overlay_directive = """

★ PRODUCT-OVERLAY MODE ACTIVE ★
The actual product image will be composited onto the scene afterwards. Therefore:
  - DO NOT render any cooler, appliance, or branded product in the scene
  - DESIGN the scene with intentional NEGATIVE SPACE in the foreground-center or right-third where the product PNG will land (about 55% of the scene height)
  - Fill the rest of the scene with people, environment, mood, lighting, props
  - The scene must work even with the product overlay removed
  - Top-left corner: keep a low-contrast clean area for logo placement"""

    user_prompt = f"""{brand_block}

POST DETAILS
Content type: {post['content_type']}
Main topic: {post['main_topic']}
Key angle: {post['key_topic']}
Audience for this post: {post['audience']}
Caption excerpt: "{post.get('caption', '')[:400]}"
Visual note from copywriter: "{post.get('visual_note', '')[:300]}"
{slide_block}{overlay_directive}

Now craft the image prompt. Return ONLY JSON:
{{"image_prompt": "...", "style_anchor": "12-word style descriptor for consistency"}}"""

    if llm_client:
        try:
            r = llm_client.generate(ART_DIRECTOR_SYSTEM, user_prompt,
                                     log_callback=log, temperature=0.75)
            prompt = (r.get("image_prompt") or "").strip()
            anchor = (r.get("style_anchor") or "").strip()
            if prompt:
                return prompt, anchor
        except Exception as e:
            if log:
                log(f"     ⚠ LLM prompt-craft failed ({e}). Using fallback.")

    # Rule-based fallback
    colors = brand.get("colors", {})
    fallback = (
        f"Premium editorial photograph for {brand.get('name', '')} "
        f"({brand.get('category', '')}). "
        f"Subject: {post['main_topic']}. {post.get('visual_note', '')}. "
        f"Color palette anchored on {colors.get('primary', '#000000')} and "
        f"{colors.get('accent', '#FFFFFF')}. Cinematic lighting, shallow depth of field, "
        f"high detail, magazine-grade composition. No text, no typography, no watermarks."
    )
    return fallback, ""


# ───────────────────────── Excel column map (matches core output) ─────────────────────────
COL = {"date": 1, "day": 2, "platform": 3, "content_type": 4, "main_topic": 5,
       "key_topic": 6, "audience": 7, "festival": 8, "hook1": 9, "hook2": 10,
       "hook3": 11, "script_slides": 12, "caption": 13, "keywords": 14,
       "hashtags": 15, "cta": 16, "visual_note": 17}


def parse_carousel_slides(script_text: str, max_slides: int = 10) -> list[dict]:
    """Split the 'Script/Slides' cell into per-slide blocks."""
    if not script_text:
        return []
    parts = re.split(r"SLIDE\s+(\d+)[:.\s]", str(script_text), flags=re.I)
    slides = []
    for i in range(1, len(parts), 2):
        try:
            n = int(parts[i])
            content = parts[i + 1].strip()
            lines = [l.strip() for l in content.split("\n") if l.strip()]
            if not lines:
                continue
            headline = lines[0]
            # If headline contains a separator, take the first half
            for sep in [" – ", " — ", " - ", " : ", ":"]:
                if sep in headline:
                    headline = headline.split(sep)[0].strip()
                    break
            body = "\n".join(lines[1:])[:500] if len(lines) > 1 else ""
            slides.append({"n": n, "headline": headline, "body": body})
        except (ValueError, IndexError):
            continue
    return slides[:max_slides]


def _slugify(text: str, max_len: int = 50) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(text or "")).strip("_")
    return (s[:max_len] or "untitled").strip("_")


# ───────────────────────── Main entry ─────────────────────────
def design_calendar_visuals(
    output_xlsx_path: str,
    brand_key: str,
    output_dir: str,
    settings: dict,
    engine: str = None,
    progress_callback=None,
    log_callback=None,
    pause_event=None,
    stop_event=None,
) -> str:
    """
    Run the Designer Agent against a completed copy Excel.
    Returns the path to the visuals folder.
    """
    freepik_key = settings.get("freepik_api_key", "").strip()
    if not freepik_key:
        raise ValueError("Freepik API key not set in Settings.")

    engine = (engine or settings.get("freepik_engine", "mystic")).lower()

    brand = core.load_brand(brand_key)
    fp = freepik_client.FreepikClient(freepik_key)

    # Load brand assets (logos + products) for compositing
    brand_assets = compositor.get_brand_assets(brand, core.BRANDS_DIR)
    has_logos = any(brand_assets["logos"].values())
    has_products = bool(brand_assets["products"])
    if log_callback:
        log_callback(f"Brand assets: logos={'✓' if has_logos else '✗'}  products={len(brand_assets['products'])}")

    try:
        llm = core.make_llm_client()
        if log_callback:
            log_callback(f"Art Director LLM: {llm.name.upper()} ({llm.model_name})")
    except Exception as e:
        if log_callback:
            log_callback(f"⚠ LLM unavailable ({e}). Using fallback prompts.")
        llm = None

    wb = load_workbook(output_xlsx_path)
    ws = wb.active

    # Read all rows into post structs
    posts = []
    for r in range(3, ws.max_row + 1):
        ct = ws.cell(r, COL["content_type"]).value
        if not ct:
            continue
        fmt = detect_content_format(str(ct))
        if fmt == "video":
            if log_callback:
                log_callback(f"  ⏭ row {r} (video): {ws.cell(r, COL['main_topic']).value}")
            continue
        posts.append({
            "row": r,
            "date": str(ws.cell(r, COL["date"]).value or ""),
            "day": str(ws.cell(r, COL["day"]).value or ""),
            "content_type": str(ct),
            "main_topic": str(ws.cell(r, COL["main_topic"]).value or ""),
            "key_topic": str(ws.cell(r, COL["key_topic"]).value or ""),
            "audience": str(ws.cell(r, COL["audience"]).value or ""),
            "festival": str(ws.cell(r, COL["festival"]).value or ""),
            "visual_note": str(ws.cell(r, COL["visual_note"]).value or ""),
            "script_slides": str(ws.cell(r, COL["script_slides"]).value or ""),
            "caption": str(ws.cell(r, COL["caption"]).value or ""),
            "format": fmt,
        })

    if not posts:
        raise ValueError("No non-video posts found in this Excel.")

    # Count total images
    total = 0
    for p in posts:
        if p["format"] == "carousel":
            slides = parse_carousel_slides(p["script_slides"])
            p["slides"] = slides or [{"n": 1, "headline": p["main_topic"], "body": ""}]
            total += len(p["slides"])
        else:
            total += 1

    safe_brand = _slugify(brand.get("name", brand_key), 30)
    visuals_root = Path(output_dir) / "visuals" / f"{safe_brand}_{datetime.now().strftime('%Y%m%d_%H%M')}"
    visuals_root.mkdir(parents=True, exist_ok=True)

    if log_callback:
        log_callback(f"\n▸ Posts to design: {len(posts)}  ·  Total images: {total}")
        log_callback(f"▸ Engine: Freepik {engine.upper()}")
        log_callback(f"▸ Visuals folder: {visuals_root}\n")

    img_count = 0
    rendered = 0
    failed_count = 0
    consecutive_failures = 0

    def _generate_one(prompt: str, aspect: str) -> str:
        if engine == "imagen3":
            return fp.generate_imagen3(prompt, aspect_ratio=aspect, log=log_callback)
        return fp.generate_mystic(prompt, aspect_ratio=aspect, log=log_callback)

    for p in posts:
        if stop_event is not None and stop_event.is_set():
            if log_callback:
                log_callback("⏹ Stopped by user.")
            break
        if pause_event is not None and pause_event.is_set():
            if log_callback:
                log_callback("⏸ Paused.")
            while pause_event.is_set():
                if stop_event is not None and stop_event.is_set():
                    break
                time.sleep(0.4)
            if log_callback:
                log_callback("▶ Resumed.")

        date_slug = _slugify(p["date"], 10)
        topic_slug = _slugify(p["main_topic"], 50)
        post_dir = visuals_root / f"{date_slug}_{topic_slug}"
        post_dir.mkdir(parents=True, exist_ok=True)

        if p["format"] == "carousel":
            style_anchor = ""
            # Detect product once for entire carousel (consistency)
            slide_product_path = compositor.detect_product(p, brand_assets["products"]) if has_products else None
            product_overlay = bool(slide_product_path)
            for slide in p["slides"]:
                img_count += 1
                label = f"Slide {slide['n']}: {p['main_topic'][:35]}"
                if progress_callback:
                    progress_callback(img_count, total, label)
                if log_callback:
                    log_callback(f"[{img_count}/{total}] Designing {label}…")
                try:
                    prompt, anchor = craft_image_prompt(
                        brand, p, slide_data=slide,
                        carousel_style_anchor=style_anchor,
                        product_overlay_mode=product_overlay,
                        llm_client=llm, log=log_callback)
                    if slide["n"] == 1 and anchor:
                        style_anchor = anchor
                    image_src = _generate_one(prompt, "square_1_1")
                    out_file = post_dir / f"slide_{slide['n']:02d}.png"
                    fp.download_to_file(image_src, str(out_file))
                    # Composite real logo + product
                    if has_logos or slide_product_path:
                        logo_path = compositor.pick_logo_variant(brand_assets["logos"], str(out_file))
                        compositor.composite_assets(
                            scene_path=str(out_file),
                            output_path=str(out_file),
                            logo_path=logo_path,
                            product_path=slide_product_path,
                            log=log_callback)
                    rendered += 1
                    consecutive_failures = 0
                    if log_callback:
                        log_callback(f"     ✓ {out_file.name}")
                except Exception as e:
                    failed_count += 1
                    consecutive_failures += 1
                    if log_callback:
                        log_callback(f"     ⚠ Failed: {e}")
                    if consecutive_failures >= 3:
                        if log_callback:
                            log_callback(f"\n✗ 3 consecutive failures — aborting. Check Freepik credits / key.")
                        return str(visuals_root)
                time.sleep(2)  # polite pacing
        else:
            img_count += 1
            label = f"{p['format'].title()}: {p['main_topic'][:40]}"
            if progress_callback:
                progress_callback(img_count, total, label)
            if log_callback:
                log_callback(f"[{img_count}/{total}] Designing {label}…")
            try:
                product_path = compositor.detect_product(p, brand_assets["products"]) if has_products else None
                product_overlay = bool(product_path)
                prompt, _ = craft_image_prompt(brand, p,
                                                 product_overlay_mode=product_overlay,
                                                 llm_client=llm, log=log_callback)
                aspect = "social_story_9_16" if p["format"] == "story" else "square_1_1"
                image_src = _generate_one(prompt, aspect)
                out_file = post_dir / f"{p['format']}.png"
                fp.download_to_file(image_src, str(out_file))
                # Composite real logo + product
                if has_logos or product_path:
                    logo_path = compositor.pick_logo_variant(brand_assets["logos"], str(out_file))
                    compositor.composite_assets(
                        scene_path=str(out_file),
                        output_path=str(out_file),
                        logo_path=logo_path,
                        product_path=product_path,
                        log=log_callback)
                rendered += 1
                consecutive_failures = 0
                if log_callback:
                    log_callback(f"     ✓ {out_file.name}")
            except Exception as e:
                failed_count += 1
                consecutive_failures += 1
                if log_callback:
                    log_callback(f"     ⚠ Failed: {e}")
                if consecutive_failures >= 3:
                    if log_callback:
                        log_callback(f"\n✗ 3 consecutive failures — aborting. Check Freepik credits / key.")
                    return str(visuals_root)
            time.sleep(2)

    if log_callback:
        log_callback(f"\n✓ Designer done — rendered {rendered}, failed {failed_count}")
        log_callback(f"  Folder: {visuals_root}")
    return str(visuals_root)
