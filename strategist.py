"""
Strategist Agent — builds a 30-day content calendar from scratch.

Inputs:
  • Brand profile (tone, audience, products)
  • Competitor playbook (what's saturated, what's underused)
  • Trend snapshot (viral moments, news, search trends)
  • Month + post count + optional theme

Output:
  • Excel calendar with: Date, Day, Platform, Content Type, Main Topic,
    Key Topic, Audience, Festival/Occasion, Notes  (no copy yet)
"""
import calendar as cal_mod
import json
import re
from datetime import date, timedelta, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


STRATEGIST_SYSTEM = """You are an elite social-media strategist who plans content for performance, not for filler. You think like a Director of Social at a top agency: every post must earn its slot. You produce calendars that performance-driven brand teams ship — NOT generic AI dumps.

You will be given:
  • The brand's full identity profile (read EVERY field — personas, cultural context, content references, content rules)
  • A competitor playbook (what's saturated vs underused in this category right now)
  • A live trends snapshot (Google search trends, Reddit virality, news headlines)
  • The target month + number of posts + optional theme
  • [If existing brand] An Analyst Baba report with real account metrics

Your job: design a 30-DAY CONTENT CALENDAR — JUST the topic layer, no copy yet. The downstream copywriter will fill in copy from your topics.

══════════════════════════════════════════════════════════════
🔴 HARD RULES (violate any → calendar fails)
══════════════════════════════════════════════════════════════

1. **TOPIC CATEGORY GUARDRAILS** — Read brand.content_rules.forbidden_topics CAREFULLY.
   Never propose topics outside the brand's actual product category.
   For a cooler brand: NO smart home, NO IoT, NO interior design, NO Home Decor,
   NO Earth Day / generic awareness days that don't tie to the product.
   For a migration agency: NO general travel guides, NO study-abroad philosophy,
   NO content unrelated to the visas they handle.

2. **EVERY TOPIC MUST MAP TO ONE OF**:
   • A specific brand product (with name from products_current list)
   • A specific audience pain point (from personas.objections or audience.pain_points)
   • A specific cultural moment (from cultural_context.local_references or seasonal_cues)
   • A specific competitor gap (from competitive_landscape — what they're missing)
   • A specific viral format opportunity (POV, meme, vox pop, myth-bust, comparison, transformation)

3. **FORMAT QUOTAS** (per month — match brand.content_rules.required_format_distribution exactly):
   Default if not specified:
     - 40% Reels (mix of POV, demo, comparison, vox pop, myth-bust)
     - 15% AI Reels (emotional CGI/cinematic storytelling — RESERVE for brand-defining moments)
     - 20% Carousels (educational, listicles, comparisons)
     - 15% Static Graphics (meme posts > festival posts > sale graphics)
     - 10% Stories (interactive, polls, Q&A)

4. **MINIMUM PER-WEEK VARIETY** (match brand.content_rules.minimum_per_week if present):
     - At least 1 POV reel per week
     - At least 1 comparison or myth-bust per week
     - At least 1 cultural / emotional moment per week
     - At least 1 product demo or test per week

5. **FORBIDDEN TOPIC TYPES** (ZERO TOLERANCE):
   ✗ "Modern living" / "Home transformation" / "Lifestyle upgrade" / "Premium experience"
   ✗ Generic awareness days unless brand directly serves that cause
   ✗ "Customer support showcase" / "We care about you" content (boring, no views)
   ✗ Smart home / IoT / connected devices for a non-IoT brand
   ✗ Home decor / styling for a non-decor brand
   ✗ Stiff corporate "Mishikas Innovates" type posts
   ✗ Long-winded "About us" topics

6. **NOTES FIELD IS GOLD** — for each post you MUST include in Notes:
   - The scroll-stop hook angle (just describe the opening, not full copy)
   - Which viral format applies (POV / Meme / Vox Pop / Myth-Bust / Comparison / Transformation / Educational / UGC / Trend-Jack)
   - The persona this targets (by name from brand.audience_personas if available)
   - The cultural/regional reference if one applies
   - The trend it rides (if trendjacking) — name the specific trend from snapshot

7. **NEVER HALLUCINATE PRODUCTS or SERVICES** — use only brand.products_current.
   If brand sells "air coolers", never write "air conditioner" / "AC" / "compressor".

8. **CONTENT REFERENCES** — if brand.content_references is populated, EXPLICITLY emulate:
   - format_winners_to_emulate patterns (POV reels, comparison reels, etc.)
   - voice_inspiration_posts vibe (e.g., "Cred-confident", "Swiggy-meme-fluent")
   - AVOID anything matching format_losers_to_avoid

9. **CULTURAL TIMING** — match seasonal_cues:
   - Peak summer month (May/June for India): heat survival + urgency dominate
   - Festival month: tie-in posts only for festivals in cultural_context.local_references
   - Off-season: brand-building, dealer content, education

10. **STRATEGIC DIFFERENTIATION** — never repeat what competitive_landscape.direct_competitors do well.
    Always lean into competitive_landscape.aspirational_reference_brands' style.
    Reference what underused_angles say in competitor playbook.

OUTPUT FORMAT — STRICT JSON, no preamble, no markdown:

{
  "calendar_meta": {
    "brand_name": "...",
    "month": "Month YYYY",
    "post_count": N,
    "recurring_series": "Name + brief description of any recurring series this month",
    "key_themes": ["3-5 thematic pillars driving the month"]
  },
  "posts": [
    {
      "date": "DD Mon",
      "day": "Day name",
      "platform": "Instagram, Facebook",
      "content_type": "Reel | Carousel | Static Graphic | Story | AI Reel",
      "main_topic": "Specific, evocative one-line idea",
      "key_topic": "The specific angle, sub-focus, or proof point",
      "audience": "Specific audience segment for this post",
      "festival": "Festival/occasion if any, else empty string",
      "notes": "1-3 sentences: tone direction, trend reference (if trendjacking, name the trend), story angle, what to highlight, what to avoid. This goes to the copywriter agent."
    }
    // ... N posts total
  ]
}

Generate the calendar now."""


# ─────────────────────────────────────────────────────────────────────────
# Helpers — month days
# ─────────────────────────────────────────────────────────────────────────
def _parse_month(month_str: str) -> tuple[int, int]:
    """Accept 'June 2026', '06/2026', '6-2026', etc. Returns (year, month)."""
    s = month_str.strip()
    # Try formats
    for fmt in ("%B %Y", "%b %Y", "%m/%Y", "%m-%Y", "%Y-%m"):
        try:
            d = datetime.strptime(s, fmt)
            return d.year, d.month
        except ValueError:
            continue
    # Fallback: split
    parts = re.split(r"[\s/\-]+", s)
    if len(parts) == 2:
        try:
            month = int(parts[0])
            year = int(parts[1])
            return year, month
        except Exception:
            pass
    raise ValueError(f"Couldn't parse month: {month_str!r}. Use e.g. 'June 2026'.")


def month_days_table(month_str: str) -> str:
    """Generate a date+day table for the strategist to reference."""
    year, month = _parse_month(month_str)
    _, last = cal_mod.monthrange(year, month)
    lines = []
    for d in range(1, last + 1):
        dt = date(year, month, d)
        lines.append(f"  {dt.strftime('%d %b')} → {dt.strftime('%A')}")
    return f"Days in {dt.strftime('%B %Y')}:\n" + "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────
# Main entry
# ─────────────────────────────────────────────────────────────────────────
def _build_strategist_brand_context(brand: dict) -> str:
    """Compose the brand context section that goes into Strategist prompt.
    Pulls from the new rich schema (personas, cultural_context, content_rules, etc.)."""
    parts = []
    parts.append(f"BRAND: {brand.get('name','')} · {brand.get('category','')}")
    parts.append(f"Tagline: {brand.get('tagline','')}")
    parts.append(f"Location: {brand.get('location','')}")

    # Business context
    biz = brand.get("business_context", {}) or {}
    if biz:
        parts.append("")
        parts.append("BUSINESS:")
        if biz.get("unique_value_proposition"):
            parts.append(f"  USP: {biz['unique_value_proposition']}")
        if biz.get("mission"):
            parts.append(f"  Mission: {biz['mission']}")
        if biz.get("pricing_strategy"):
            parts.append(f"  Pricing: {biz['pricing_strategy']}")

    # Products (the only legal subjects for product-led posts)
    products = brand.get("products_current") or brand.get("products_or_services") or []
    if products:
        parts.append("")
        parts.append("ONLY THESE PRODUCTS EXIST (use exact names):")
        for p in products[:8]:
            parts.append(f"  • {p}")
    not_made = brand.get("products_NOT_made", [])
    if not_made:
        parts.append("DOES NOT MAKE (forbidden mentions):")
        for n in not_made:
            parts.append(f"  ✗ {n}")

    # Personas
    personas = brand.get("audience_personas", []) or []
    if personas:
        parts.append("")
        parts.append("AUDIENCE PERSONAS (target individual personas, not 'everyone'):")
        for p in personas[:3]:
            parts.append(f"  • {p.get('name','')} ({p.get('age','')}) — {p.get('role','')}")
            if p.get("decision_triggers"):
                parts.append(f"      triggers: {p['decision_triggers'][:200]}")
            if p.get("objections"):
                parts.append(f"      objections: {p['objections'][:200]}")
    elif brand.get("audience", {}).get("primary"):
        parts.append("")
        parts.append(f"AUDIENCE: {brand['audience']['primary'][:300]}")

    # Cultural context
    cc = brand.get("cultural_context", {}) or {}
    if cc:
        parts.append("")
        parts.append("CULTURAL CONTEXT:")
        if cc.get("region"):
            parts.append(f"  Region: {cc['region']}")
        if cc.get("climate_reality"):
            parts.append(f"  Climate: {cc['climate_reality']}")
        if cc.get("seasonal_cues"):
            parts.append("  Seasonal cues:")
            for k, v in list(cc["seasonal_cues"].items())[:6]:
                parts.append(f"    {k}: {str(v)[:120]}")
        if cc.get("local_references"):
            parts.append(f"  Local references (use specifically): "
                        + ", ".join(cc["local_references"][:8]))
        if cc.get("no_go_topics"):
            parts.append(f"  NO-GO topics: {', '.join(cc['no_go_topics'][:6])}")

    # Competitive landscape
    comp = brand.get("competitive_landscape", {}) or {}
    if comp:
        parts.append("")
        parts.append("COMPETITIVE POSITIONING:")
        if comp.get("one_line_differentiation"):
            parts.append(f"  Differentiation: {comp['one_line_differentiation']}")
        if comp.get("aspirational_reference_brands"):
            parts.append("  Aspirational vibes (emulate THESE):")
            for b in comp["aspirational_reference_brands"][:5]:
                parts.append(f"    • {b}")
        if comp.get("anti_brands"):
            parts.append("  Anti-brands (do NOT sound like):")
            for b in comp["anti_brands"][:4]:
                parts.append(f"    ✗ {b}")

    # Content references — gold for tone
    refs = brand.get("content_references", {}) or {}
    if refs:
        parts.append("")
        parts.append("CONTENT REFERENCE (mimic these patterns):")
        if refs.get("format_winners_to_emulate"):
            parts.append("  Format winners:")
            for w in refs["format_winners_to_emulate"][:6]:
                parts.append(f"    ✓ {w}")
        if refs.get("format_losers_to_avoid"):
            parts.append("  AVOID these formats:")
            for w in refs["format_losers_to_avoid"][:6]:
                parts.append(f"    ✗ {w}")

    # Content rules — the hardest constraints
    rules = brand.get("content_rules", {}) or {}
    if rules:
        parts.append("")
        parts.append("CONTENT RULES (hard constraints):")
        if rules.get("approved_pillars"):
            parts.append("  Approved pillars (every post must serve one):")
            for p in rules["approved_pillars"][:6]:
                parts.append(f"    • {p}")
        if rules.get("forbidden_topics"):
            parts.append("  FORBIDDEN topics:")
            for f in rules["forbidden_topics"][:8]:
                parts.append(f"    ✗ {f}")
        if rules.get("required_format_distribution_per_month"):
            mix = ", ".join(f"{k}: {v}%" for k, v in
                          rules["required_format_distribution_per_month"].items())
            parts.append(f"  Required format mix: {mix}")
        if rules.get("minimum_per_week"):
            mw = ", ".join(f"{k}={v}" for k, v in rules["minimum_per_week"].items())
            parts.append(f"  Minimum per week: {mw}")

    # Tone & voice
    tone = brand.get("tone", {}) or {}
    if tone.get("personality"):
        parts.append("")
        parts.append(f"VOICE: {tone.get('voice','')}")
        parts.append(f"PERSONALITY: {tone.get('personality','')}")

    # Hook library
    hooks = brand.get("hook_library", []) or []
    if hooks:
        parts.append("")
        parts.append("BRAND HOOK LIBRARY (these are PROVEN tone-perfect openers — reference style):")
        for h in hooks[:10]:
            parts.append(f"  ▸ {h}")

    return "\n".join(parts)


def generate_topic_calendar(brand: dict,
                             month: str,
                             post_count: int,
                             theme: str,
                             llm_client,
                             competitor_playbook: str = "",
                             trends_brief: str = "",
                             strategy_brief: str = "",
                             log=None) -> dict:
    """Call the Strategist LLM and return the calendar JSON."""
    if log:
        log(f"▸ Strategist designing {post_count} posts for {month}…")

    days_block = month_days_table(month)
    brand_block = _build_strategist_brand_context(brand)

    # Inject hook + format references from format_templates
    try:
        import format_templates as _ft
        hook_block = _ft.hook_patterns_brief()
    except Exception:
        hook_block = ""

    strategy_block = ""
    if strategy_brief.strip():
        strategy_block = ("GROWTH PLANNER STRATEGY (your north star — follow it strictly)\n"
                          "────────────────────────────────────────────\n"
                          + strategy_brief + "\n\n")

    user_prompt = f"""BRAND CONTEXT
─────────────
{brand_block}

{hook_block}

MONTH: {month}
NUMBER OF POSTS: {post_count}
OPTIONAL THEME: {theme or 'No specific theme — balanced editorial mix per brand.content_rules.'}

{days_block}

{strategy_block}{('COMPETITOR PLAYBOOK' + chr(10) + '──────────────────' + chr(10) + competitor_playbook + chr(10) + chr(10)) if competitor_playbook else ''}{('LIVE TRENDS SNAPSHOT' + chr(10) + '────────────────────' + chr(10) + trends_brief + chr(10) + chr(10)) if trends_brief else ''}

Now design the {post_count}-post calendar for {month}.

⚠ REMINDER OF YOUR HARD RULES:
  • Every topic must hit one of: product / pain / culture / competitor-gap / viral-format
  • No off-category topics (no Home Decor / Smart Home / Earth Day for a cooler brand)
  • Format mix must match brand.content_rules.required_format_distribution
  • Notes field MUST include: hook angle + format type + persona + cultural reference
  • Lean into brand's hook_library style — never brochure language

Return ONLY the JSON object."""

    out = llm_client.generate(STRATEGIST_SYSTEM, user_prompt, log_callback=log,
                                temperature=0.85, max_tokens=6000)
    if log:
        log(f"✓ Strategist produced {len(out.get('posts', []))} topics.")
    return out


# ─────────────────────────────────────────────────────────────────────────
# Excel writer for topic calendar (just the input layer)
# ─────────────────────────────────────────────────────────────────────────
HEADERS = ["Date", "Day", "Platform", "Content Type", "Main Topic",
           "Key Topic", "Audience", "Festival / Occasion", "Notes"]
WIDTHS = [12, 12, 22, 16, 38, 38, 28, 22, 50]

CONTENT_TYPE_COLORS = {
    "AI Reel":         "FFE7DF",
    "Reel":            "E6EEF8",
    "Carousel":        "EFEDFB",
    "Static Graphic":  "FFF8E1",
    "Story":           "E8F5EC",
}


def write_topic_calendar_xlsx(brand_name: str, month: str, calendar_data: dict,
                               output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Content Calendar"

    title = f"{brand_name.upper()} — STRATEGIST CONTENT CALENDAR ({month})"
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    c = ws.cell(row=1, column=1)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003995")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    meta = calendar_data.get("calendar_meta", {})
    sub = ""
    if meta.get("key_themes"):
        sub += "Themes: " + ", ".join(meta["key_themes"][:4])
    if meta.get("recurring_series"):
        sub += f"   ·   Recurring: {meta['recurring_series']}"
    ws.cell(row=2, column=1, value=sub or "Auto-generated by Strategist agent")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    s = ws.cell(row=2, column=1)
    s.font = Font(italic=True, size=10, color="555555")
    s.fill = PatternFill("solid", fgColor="F8EBE0")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor="0D1B3D")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 35

    thin = Side(border_style="thin", color="D5D5D5")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for i, p in enumerate(calendar_data.get("posts", []), start=4):
        ctype = p.get("content_type", "")
        row_fill = CONTENT_TYPE_COLORS.get(ctype, "FFFFFF")
        values = [
            p.get("date", ""), p.get("day", ""), p.get("platform", "Instagram, Facebook"),
            ctype, p.get("main_topic", ""), p.get("key_topic", ""),
            p.get("audience", ""), p.get("festival", ""), p.get("notes", ""),
        ]
        for c_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=i, column=c_idx, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            cell.font = Font(size=10)
            if c_idx == 4:
                cell.fill = PatternFill("solid", fgColor=row_fill)
                cell.font = Font(size=10, bold=True, color="0D1B3D")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFAFA")
        ws.row_dimensions[i].height = 75

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    wb.save(output_path)
    return output_path
