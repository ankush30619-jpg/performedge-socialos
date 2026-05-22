"""Generate empty input calendar templates for both brands."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent

HEADERS = ["Date", "Day", "Platform", "Content Type", "Main Topic",
           "Key Topic", "Audience", "Festival / Occasion", "Notes"]

WIDTHS = [14, 12, 22, 16, 36, 28, 26, 22, 32]


def make_template(brand_name: str, brand_color_hex: str, sample_rows: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Content Calendar"

    title = f"{brand_name.upper()} — CONTENT CALENDAR (Input Template)"
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    c = ws.cell(row=1, column=1)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=brand_color_hex.lstrip("#"))
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D1B3D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32

    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for r_idx, row_data in enumerate(sample_rows, start=3):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
        ws.row_dimensions[r_idx].height = 50

    # Add a few empty rows
    for r in range(3 + len(sample_rows), 3 + len(sample_rows) + 12):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 50

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    return wb


MISHIKA_SAMPLES = [
    ["10 May", "Saturday", "Instagram, Facebook", "Reel",
     "Brand Launch / Intro — Who is Mishikas Electronics?",
     "Hinglish presenter introduces brand and product range",
     "Family buyers 30-45, North India",
     "Brand Launch Day",
     "Energetic, fast-cuts, trending BGM. Showcase upcoming products too."],
    ["12 May", "Monday", "Instagram, Facebook", "Static Graphic",
     "Mother's Day — Gift Maa a Smart Home",
     "Emotional graphic — Maa ke liye smart home from Mishikas",
     "Adult sons/daughters buying gifts for mother",
     "Mother's Day",
     "Warm gradient, illustration style, festive badge."],
    ["13 May", "Tuesday", "Instagram, Facebook", "Carousel",
     "Top 5 Home Appliances Under ₹5000",
     "Swipeable carousel — 1 product per slide with price + 3 features",
     "Budget-conscious family buyers",
     "",
     "7 slides. Slide 1 cover, slides 2-6 products, slide 7 CTA."],
]

JOURNEY_SAMPLES = [
    ["01 May", "Thursday", "Instagram, Facebook", "Carousel",
     "Student Visa (500) — Complete Guide: What You CAN & CANNOT Do",
     "Work rights, OSHC, financial requirement, extension rules",
     "Student Visa 500 holders",
     "",
     "Use verified facts from journeygrp.com. Quote specific numbers."],
    ["02 May", "Friday", "Instagram, Facebook, YouTube", "Reel",
     "485 Visa — Two Streams Explained (PSWS vs Graduate Work)",
     "Eligibility, duration, fees, age limit, work rights",
     "485 visa holders / soon-to-be graduates",
     "",
     "On-camera explainer style. Australian English. Reference DoHA."],
    ["07 May", "Wednesday", "Instagram, Facebook", "Static Post",
     "Visa Grant Announcement — [Subclass to fill]",
     "Real client visa grant — quote review, agent credit",
     "All migrants",
     "",
     "Use Maheep Singh Virk as agent. MARA registered. Real client quote."],
]


if __name__ == "__main__":
    out_mishika = ROOT / "Mishikas_Calendar_Template.xlsx"
    wb1 = make_template("Mishikas Electronics", "#003995", MISHIKA_SAMPLES)
    wb1.save(out_mishika)
    print(f"Created: {out_mishika}")

    out_journey = ROOT / "JourneyGroup_Calendar_Template.xlsx"
    wb2 = make_template("Journey Group", "#0D1B3D", JOURNEY_SAMPLES)
    wb2.save(out_journey)
    print(f"Created: {out_journey}")
