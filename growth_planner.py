"""
Growth Planner Agent — replaces the old Trend Scout with a much bigger role.

What it does:
  1. Scrapes the brand's website to deeply understand the business
  2. Runs deep web research via Tavily (industry strategies, what's working in 2026)
  3. Harvests live trends (Google Trends + Reddit + NewsAPI)
  4. Reads ongoing campaign notes + reference PDFs/images
  5. Feeds everything to an LLM acting as senior growth strategist
  6. Produces:
       a) Strategy JSON   (machine-readable, feeds Strategist agent)
       b) Strategy PPTX   (polished branded deck for client review)
"""
import io
import json
import re
import time
import requests
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse, urljoin

from bs4 import BeautifulSoup

import core
import trend_scout


# ─────────────────────────────────────────────────────────────────────────
# Web scraping
# ─────────────────────────────────────────────────────────────────────────
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 SocialCopyStudio/3.0"

KEY_PATH_HINTS = ["about", "products", "services", "what-we-do", "our-story",
                   "company", "brand", "story", "team", "why", "category"]


def _clean_html_text(html: str, max_chars: int = 8000) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "iframe", "svg",
                      "nav", "footer", "header", "form"]):
        tag.decompose()
    text = soup.get_text(separator=" ", strip=True)
    text = re.sub(r"\s+", " ", text)
    return text[:max_chars]


def _fetch(url: str, timeout: int = 20) -> str | None:
    try:
        r = requests.get(url, headers={"User-Agent": USER_AGENT},
                          timeout=timeout, allow_redirects=True)
        if r.status_code == 200 and "text/html" in r.headers.get("Content-Type", ""):
            return r.text
    except Exception:
        return None
    return None


def scrape_brand_website(url: str, log=None) -> dict:
    """
    Pull main content from a brand website — homepage + best matched
    about/products/services pages.
    Returns {"homepage": "...text...", "pages": {url: text, ...}, "domain": "..."}
    """
    if not url or not url.startswith(("http://", "https://")):
        url = "https://" + (url or "").lstrip("/")

    parsed = urlparse(url)
    domain = parsed.netloc
    if log:
        log(f"  ⟳ Scraping {domain}…")

    out = {"homepage": "", "pages": {}, "domain": domain, "url": url}
    home_html = _fetch(url)
    if not home_html:
        if log:
            log(f"     ⚠ Couldn't fetch homepage.")
        return out
    out["homepage"] = _clean_html_text(home_html)

    # Find internal links matching key hints
    soup = BeautifulSoup(home_html, "html.parser")
    candidates = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
            continue
        full = urljoin(url, href)
        parsed_link = urlparse(full)
        if parsed_link.netloc and parsed_link.netloc != domain:
            continue
        path = parsed_link.path.lower()
        for hint in KEY_PATH_HINTS:
            if hint in path:
                candidates.append(full)
                break

    # Dedupe + cap
    seen = set()
    unique = []
    for c in candidates:
        if c not in seen and len(unique) < 5:
            seen.add(c); unique.append(c)

    for u in unique:
        html = _fetch(u)
        if not html:
            continue
        out["pages"][u] = _clean_html_text(html, max_chars=4000)
        if log:
            log(f"     ✓ {urlparse(u).path}")
        time.sleep(0.5)

    if log:
        log(f"     ✓ Scraped {1 + len(out['pages'])} page(s), ~{len(out['homepage']) + sum(len(v) for v in out['pages'].values())} chars.")
    return out


# ─────────────────────────────────────────────────────────────────────────
# Self social media analysis (the brand's own IG / FB)
# ─────────────────────────────────────────────────────────────────────────
def analyse_self_social(brand_key: str, brand: dict, ig_handle: str,
                         fb_handle: str, settings: dict, llm_client,
                         log=None) -> dict:
    """
    Scrape the brand's OWN Instagram (via Apify) and produce a baseline
    snapshot of current state: followers, engagement, content themes, top posts.
    """
    import competitor_analyzer

    result = {
        "instagram_handle": ig_handle or "",
        "facebook_handle": fb_handle or "",
        "ig_posts_count": 0,
        "ig_followers": None,
        "ig_following": None,
        "ig_bio": "",
        "ig_recent_posts": [],
        "engagement_summary": "",
        "content_themes": [],
        "what_is_working": [],
        "gaps": [],
    }

    apify_token = settings.get("apify_token", "")
    if not ig_handle or not apify_token:
        if log:
            log("  ⚠ Skipping self-IG scrape (handle or Apify token missing).")
        return result

    handle = competitor_analyzer._normalise_handle(ig_handle)
    if log:
        log(f"  ⟳ Scraping own Instagram @{handle}…")
    try:
        posts = competitor_analyzer.scrape_instagram_profile(
            handle, apify_token, limit=20, log=log)
    except Exception as e:
        if log:
            log(f"  ⚠ Self IG scrape failed: {e}")
        return result

    if not posts:
        if log:
            log("  ⚠ No posts returned — profile may be empty/private/new.")
        return result

    # Extract profile-level stats from first post (Apify schema)
    first = posts[0] if posts else {}
    owner = first.get("ownerUsername", handle)
    # Apify Instagram scraper returns ownerFullName, followersCount fields when available
    result["ig_followers"] = first.get("ownerFollowersCount") or first.get("followersCount")
    result["ig_following"] = first.get("ownerFollowingCount") or first.get("followingCount")
    result["ig_bio"] = first.get("ownerBio") or first.get("biography", "")
    result["ig_posts_count"] = first.get("ownerPostsCount") or len(posts)

    # Recent posts summary
    recent = []
    total_likes = 0
    total_comments = 0
    for p in posts[:15]:
        recent.append({
            "type": p.get("type") or p.get("productType", ""),
            "caption": (p.get("caption") or "")[:300],
            "hashtags": (p.get("hashtags") or [])[:10],
            "likes": p.get("likesCount", 0),
            "comments": p.get("commentsCount", 0),
            "timestamp": p.get("timestamp", ""),
        })
        total_likes += p.get("likesCount", 0) or 0
        total_comments += p.get("commentsCount", 0) or 0
    result["ig_recent_posts"] = recent

    # Engagement rate estimate
    avg_engagement = (total_likes + total_comments) / max(len(recent), 1)
    er = None
    if result["ig_followers"] and result["ig_followers"] > 0:
        er = (avg_engagement / result["ig_followers"]) * 100
        result["engagement_summary"] = (
            f"~{int(avg_engagement)} avg interactions per post "
            f"on {result['ig_followers']} followers → "
            f"engagement rate ≈ {er:.2f}%")
    else:
        result["engagement_summary"] = f"~{int(avg_engagement)} avg interactions per post (followers count not available)"

    if log:
        log(f"  ✓ Own IG: {len(recent)} posts, {result['engagement_summary']}")

    # LLM analysis — themes, what works, gaps
    if llm_client:
        try:
            if log:
                log("  ⟳ Analysing own social content…")
            analysis_prompt = f"""You are auditing this brand's OWN Instagram. Identify:
1. Top 3-5 content themes they're already posting about
2. What's WORKING (highest engagement themes, format patterns)
3. Gaps / weaknesses (content types they're NOT trying, low engagement themes)

OWN POSTS (most recent first):
{json.dumps(recent, indent=2, ensure_ascii=False)[:8000]}

ENGAGEMENT SUMMARY: {result['engagement_summary']}
TOTAL FOLLOWERS: {result['ig_followers'] or 'unknown'}

Return JSON: {{
  "content_themes": ["..."],
  "what_is_working": ["..."],
  "gaps": ["..."],
  "honest_assessment": "1-2 sentences on the brand's current social presence"
}}"""
            r = llm_client.generate(
                "You are a senior social media auditor.",
                analysis_prompt, log_callback=log, temperature=0.4)
            result["content_themes"] = r.get("content_themes", [])
            result["what_is_working"] = r.get("what_is_working", [])
            result["gaps"] = r.get("gaps", [])
            result["honest_assessment"] = r.get("honest_assessment", "")
        except Exception as e:
            if log:
                log(f"  ⚠ Self-analysis LLM call failed: {e}")
    return result


def format_self_social_brief(self_data: dict) -> str:
    """Compact summary for Growth Planner LLM prompt."""
    if not self_data or not self_data.get("instagram_handle"):
        return ""
    parts = [f"OWN INSTAGRAM: @{self_data['instagram_handle']}"]
    if self_data.get("ig_followers") is not None:
        parts.append(f"Current followers: {self_data['ig_followers']}")
    if self_data.get("engagement_summary"):
        parts.append(self_data["engagement_summary"])
    if self_data.get("ig_bio"):
        parts.append(f"Bio: {self_data['ig_bio'][:200]}")
    if self_data.get("honest_assessment"):
        parts.append(f"\nHonest assessment: {self_data['honest_assessment']}")
    if self_data.get("content_themes"):
        parts.append("\nContent themes already posted:")
        parts.extend(f"  • {x}" for x in self_data["content_themes"])
    if self_data.get("what_is_working"):
        parts.append("\nWhat seems to be working:")
        parts.extend(f"  • {x}" for x in self_data["what_is_working"])
    if self_data.get("gaps"):
        parts.append("\nGaps / opportunities on own page:")
        parts.extend(f"  • {x}" for x in self_data["gaps"])
    if self_data.get("facebook_handle"):
        parts.append(f"\nFacebook page: {self_data['facebook_handle']}")
    return "\n".join(parts)


# ─────────────────────────────────────────────────────────────────────────
# Tavily web research
# ─────────────────────────────────────────────────────────────────────────
def tavily_research(brand: dict, tavily_key: str, log=None) -> dict:
    """Run a series of strategic queries via Tavily. Returns aggregated results."""
    if not tavily_key:
        if log:
            log("  ⚠ Tavily key not set — skipping web research.")
        return {"queries": [], "results": []}

    try:
        from tavily import TavilyClient
    except ImportError:
        if log:
            log("  ⚠ Tavily SDK not installed.")
        return {"queries": [], "results": []}

    client = TavilyClient(api_key=tavily_key)
    name = brand.get("name", "")
    category = brand.get("category", "")
    location = brand.get("location", "India")
    country = "India" if "india" in location.lower() else (
              "Australia" if "australia" in location.lower() else "global")

    queries = [
        f"{category} Instagram growth strategy 2026 {country}",
        f"best performing social media content for {category}",
        f"viral Instagram trends for {category} brands 2026",
        f"how {category} brands grow followers on Instagram and Facebook",
        f"top {category} brands social media case study",
    ]

    aggregated = {"queries": queries, "results": []}
    for q in queries:
        try:
            if log:
                log(f"  ⟳ Tavily: {q[:60]}…")
            r = client.search(query=q, max_results=5, search_depth="advanced",
                              include_answer=True)
            answer = r.get("answer", "")
            results = []
            for x in r.get("results", []):
                results.append({
                    "title": x.get("title", "")[:200],
                    "url": x.get("url", ""),
                    "content": x.get("content", "")[:1200],
                    "score": x.get("score", 0),
                })
            aggregated["results"].append({"query": q, "answer": answer, "results": results})
            time.sleep(0.8)  # be polite
        except Exception as e:
            if log:
                log(f"     ⚠ Tavily query failed: {e}")
    if log:
        log(f"  ✓ Tavily: {sum(len(b['results']) for b in aggregated['results'])} sources gathered.")
    return aggregated


# ─────────────────────────────────────────────────────────────────────────
# Strategy LLM call
# ─────────────────────────────────────────────────────────────────────────
GROWTH_PLANNER_SYSTEM = """You are GROOK — an elite social media growth strategist with 15+ years experience building Instagram + Facebook strategies for premium D2C and service brands across India, Australia, and global markets.

You think like a Director of Social at a top digital agency: business-first, audience-first, results-obsessed. You do NOT produce generic AI playbooks ("post products, then reviews, then UGC"). You earn every recommendation from real data.

Your output drives a 30-DAY social plan focused on TWO KPIs:
  1. Follower growth (absolute + %)
  2. Engagement rate (avg engagement / followers per post)

You operate in TWO modes — read carefully which one applies:

══════════════════════════════════════════════════════════════════
MODE 1 — EXISTING BRAND
You have Analyst Baba's report with real Meta data: follower count, ER,
top/worst posts, audience demographics, what's working, what's failing.
Plus a previous plan (if this is v2+).

In this mode your strategy MUST:
  • Quote real numbers from the audit (current followers, current ER)
  • Reference top-performing post themes by name/format
  • Calibrate KPI targets to actual baselines (don't say 25K followers if at 397)
  • Double-down on what's working, cut what isn't
  • If v2+: explicitly call out what changes from v(N-1)

══════════════════════════════════════════════════════════════════
MODE 2 — NEW BRAND
No own-account data. Brand is launching from zero or just-launched.
Build the plan from industry baselines + competitor patterns + brand identity.

In this mode your strategy MUST:
  • Set realistic month-1 KPI targets for a new account
    (typical: 200-1500 followers, focus on hitting baseline ER)
  • Front-load brand-introduction content (founder story, brand world)
  • Avoid pretending you have data you don't have
  • Lean heavily on category baselines and competitor analysis
══════════════════════════════════════════════════════════════════

Inputs you'll receive:
  • Brand profile (tone, voice, audience, products)
  • Brand mode: "existing" or "new"
  • Real text scraped from brand website
  • Ongoing campaigns
  • Tavily web research (what's working in this category in 2026)
  • Live trends snapshot (Google Trends, Reddit, news)
  • Competitor playbook (if competitor research has been run)
  • [EXISTING only] Analyst Baba's latest report (real metrics + recommendations)
  • [v2+ only] Previous plan version (so you can show evolution)

Synthesise it all and produce a complete strategic plan as a JSON object — every field rigorously filled, no generic platitudes, no consulting clichés.

OUTPUT JSON SCHEMA — return EXACTLY this shape, no preamble, no markdown:

{
  "executive_summary": {
    "current_state": "2-3 sentences on where the brand stands today on social — for EXISTING brands quote real numbers from Analyst Baba; for NEW brands say so honestly",
    "opportunity": "2-3 sentences naming the specific opportunity in the category right now",
    "north_star_metric": "The single most important metric for the next 30 days (usually Followers OR Engagement Rate)",
    "headline_target": "Specific 30-day target (e.g., '397 → 1,200 followers · 5.97% → 7%+ ER' for existing; '0 → 500 followers · hit 1.5% ER' for new)"
  },
  "business_understanding": {
    "what_brand_does": "1-2 sentences",
    "unique_value_proposition": "1-2 sentences specific to this brand",
    "primary_business_goals": ["3-4 concrete business goals for next 90 days"]
  },
  "industry_landscape": {
    "category_state_2026": "What's happening in this category on social media right now (2-3 sentences)",
    "key_trends_in_category": ["4-6 specific trends in this category"],
    "what_top_brands_do_well": ["3-5 specific tactics observed from research"],
    "what_top_brands_miss": ["2-4 gaps competitors are leaving on the table"]
  },
  "audience_segments": [
    {
      "name": "Persona name (e.g., 'Family Buyer Maa')",
      "description": "1-2 sentences describing this segment",
      "pain_points": ["3-4 specific pain points"],
      "desires": ["3-4 specific desires"],
      "content_preference": "what kind of content this persona consumes"
    }
  ],
  "growth_pillars": [
    {
      "name": "Pillar name (e.g., 'Cinematic Emotional Storytelling')",
      "thesis": "Why this pillar wins for THIS brand (1-2 sentences)",
      "content_formats": ["AI Reel", "Carousel", "..."],
      "share_of_calendar": "20-40 (a number — % of monthly posts in this pillar)",
      "example_content_themes": ["3-5 example content themes this pillar produces"]
    }
  ],
  "content_strategy": {
    "format_mix": {"Reel": 25, "AI Reel": 20, "Carousel": 25, "Static Graphic": 20, "Story": 10},
    "posting_frequency": "e.g., '5-6 feed posts per week + daily stories'",
    "hook_strategy": "1-2 sentences on hook approach for this brand",
    "hashtag_strategy": "1-2 sentences on hashtag approach (count, mix, branded vs trending)",
    "best_posting_times": "specific time windows for primary audience"
  },
  "campaign_roadmap": {
    "week_1_theme": "Theme + 1-line description",
    "week_1_campaigns": ["2-3 specific campaigns for week 1"],
    "week_2_theme": "Theme + 1-line description",
    "week_2_campaigns": ["2-3 specific campaigns"],
    "week_3_theme": "Theme + 1-line description",
    "week_3_campaigns": ["2-3 specific campaigns"],
    "week_4_theme": "Theme + 1-line description",
    "week_4_campaigns": ["2-3 specific campaigns"]
  },
  "trend_watch": [
    {
      "trend": "Specific trend name/description",
      "why_relevant": "Why it fits THIS brand",
      "how_to_ride": "Concrete content idea"
    }
  ],
  "kpis": {
    "follower_growth_target_30_days": "Specific number with rationale",
    "engagement_rate_target": "% target (be realistic given baseline)",
    "reach_target": "Specific reach goal",
    "save_share_target": "Save/share target",
    "conversion_target": "DM/lead/sale target (whatever applies to this brand)"
  },
  "recommendations_priority_order": [
    "5-8 specific recommendations the brand should act on in next 30 days, in priority order"
  ],
  "changes_from_previous_version": [
    "v2+ ONLY: list what specifically changed vs v(N-1). Include: pillar weight changes, target recalibrations, format-mix shifts, dropped/added campaigns. Cite Analyst data that drove each change."
  ]
}

QUALITY RULES — non-negotiable:
1. EVERYTHING must be specific to this brand and category. Generic = failure.
2. Use the actual brand voice and audience details. Don't recycle generic personas.
3. Numbers in KPIs must be realistic for the brand's current state (a new brand at 0 followers ≠ 200K target in 90 days).
4. Trend Watch entries must reference the actual trends snapshot provided. Don't invent trends.
5. Growth pillars must add to 100% in share_of_calendar.
6. For Indian brands, lean into Indian cultural context and Hinglish where appropriate.
7. NEVER hallucinate products the brand doesn't make. Read the brand profile carefully.
"""


def generate_strategy_json(brand: dict,
                            website_data: dict,
                            ongoing_campaigns_text: str,
                            references_text: str,
                            tavily_data: dict,
                            trends_brief: str,
                            competitor_playbook: str,
                            self_social_brief: str = "",
                            manual_analytics: str = "",
                            analyst_brief: str = "",
                            previous_plan_brief: str = "",
                            brand_mode: str = "new",
                            plan_version: int = 1,
                            llm_client=None,
                            log=None) -> dict:
    """Call the LLM strategist with all inputs, return the structured strategy."""
    # Prepare condensed inputs
    web_block = ""
    if website_data.get("homepage"):
        web_block += "HOMEPAGE TEXT:\n" + website_data["homepage"][:4500]
    for u, t in (website_data.get("pages") or {}).items():
        web_block += f"\n\nPAGE {urlparse(u).path}:\n{t[:1800]}"

    tavily_block = ""
    for batch in tavily_data.get("results", []):
        tavily_block += f"\n\n## Query: {batch['query']}\n"
        if batch.get("answer"):
            tavily_block += f"Tavily summary: {batch['answer'][:600]}\n"
        for r in batch.get("results", [])[:3]:
            tavily_block += f"  • {r['title']}\n    {r['content'][:400]}\n"

    mode_banner = (
        f"🟢 BRAND MODE: EXISTING\n"
        f"This brand has live Meta data (Analyst Baba). Calibrate targets to real baselines."
        if brand_mode == "existing"
        else
        f"🆕 BRAND MODE: NEW\n"
        f"This brand has no live Meta data. Build the plan from industry baselines + competitor patterns + brand identity.")

    version_banner = (
        f"\n📍 PLAN VERSION: v{plan_version}"
        + (" (revising previous plan based on Analyst data)" if plan_version > 1 else " (first plan for this brand)"))

    # Trim brand JSON to the strategic essentials only
    brand_compact = {
        "name": brand.get("name"),
        "tagline": brand.get("tagline"),
        "category": brand.get("category"),
        "location": brand.get("location"),
        "products_current": brand.get("products_current") or brand.get("products_or_services"),
        "products_NOT_made": brand.get("products_NOT_made", []),
        "tone": brand.get("tone", {}),
        "audience_primary": (brand.get("audience") or {}).get("primary", ""),
        "signature_phrases": brand.get("signature_phrases", [])[:6],
    }

    parts = [
        mode_banner + version_banner,
        "",
        "BRAND PROFILE",
        "────────────",
        json.dumps(brand_compact, indent=2, ensure_ascii=False)[:2500],
        "",
        "BRAND WEBSITE EXCERPT",
        "──────────────────────",
        (web_block or "(none)")[:3500],
        "",
    ]

    if brand_mode == "existing":
        parts += [
            "🔮 ANALYST BABA REPORT — your single most important input",
            "──────────────────────────────────────────────────────────",
            analyst_brief[:3500] or "(no Analyst report yet — run Analyst Baba first for better data)",
            "",
        ]
        if plan_version > 1 and previous_plan_brief:
            parts += [
                "📋 PREVIOUS PLAN VERSION — what you said last time",
                "─────────────────────────────────────────────────",
                previous_plan_brief[:2000],
                "",
                "👉 You MUST populate 'changes_from_previous_version' with specific deltas vs v"
                + str(plan_version - 1) + ".",
                "",
            ]

    parts += [
        "BRAND'S OWN SOCIAL MEDIA (scraped)",
        "──────────────────────────────────",
        self_social_brief[:2000] or "(brand has no public IG presence yet)",
        "",
        "USER-PROVIDED ANALYTICS / NOTES",
        "───────────────────────────────",
        manual_analytics[:1500] or "(none)",
        "",
        "ONGOING CAMPAIGNS",
        "─────────────────",
        ongoing_campaigns_text[:1500] or "(none)",
        "",
        "REFERENCE MATERIALS",
        "───────────────────",
        references_text[:1500] or "(none)",
        "",
        "WEB RESEARCH (Tavily — what's working in this category)",
        "─────────────────────────────────────────────────────────",
        tavily_block[:3500] or "(none)",
        "",
        "TRENDS SNAPSHOT",
        "───────────────",
        trends_brief[:2000] or "(none)",
        "",
        "COMPETITOR PLAYBOOK",
        "───────────────────",
        competitor_playbook[:2000] or "(none — run Competitors agent for sharper differentiation)",
        "",
        "",
        "🎯 PRODUCE THE 30-DAY GROWTH PLAN JSON.",
        f"   Mode: {brand_mode.upper()}  ·  Version: v{plan_version}",
        "   KPIs locked to: follower growth + engagement rate.",
        "   Calibrate targets to the actual data shown above — no aspirational nonsense.",
        "   Return ONLY the JSON object.",
    ]
    user_prompt = "\n".join(parts)

    if log:
        log(f"  · GROOK payload: {len(user_prompt)//4} estimated tokens "
            f"({len(user_prompt)} chars)")

    if log:
        log(f"  ⟳ LLM strategist thinking… (input ≈ {len(user_prompt)//1000}k chars)")

    out = llm_client.generate(GROWTH_PLANNER_SYSTEM, user_prompt,
                                log_callback=log, temperature=0.7, max_tokens=10000)
    if log:
        log("  ✓ Strategy JSON produced.")
    return out


# ─────────────────────────────────────────────────────────────────────────
# Analytics file extractor — Excel / PDF / PNG / JPG
# ─────────────────────────────────────────────────────────────────────────
ANALYTICS_SUMMARY_PROMPT = """You are reading a brand's social media analytics report. Extract every concrete number and observation that matters for strategy:
  • Follower count + growth rate
  • Reach, impressions, profile visits, website clicks
  • Engagement rate, total likes, comments, shares, saves
  • Top-performing posts (themes, formats)
  • Worst-performing posts
  • Audience demographics (age, gender, location)
  • Best posting times / days
  • Story metrics (taps forward / back / exits)
  • Any campaign-specific results

Return ONLY JSON: {"analytics_summary": "A dense bullet-list of every concrete data point. Each bullet is one fact. 15-30 bullets ideally."}"""


def _extract_excel_text(path: Path, max_chars: int = 12000) -> str:
    """Dump every sheet of an Excel file into readable text."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""
    try:
        wb = load_workbook(str(path), data_only=True)
    except Exception as e:
        return f"[Excel parse error: {e}]"
    out = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        out.append(f"\n=== SHEET: {sn} ===")
        for row in ws.iter_rows(values_only=True):
            if any(c not in (None, "") for c in row):
                cells = [str(c) if c is not None else "" for c in row]
                out.append(" | ".join(cells))
            if sum(len(x) for x in out) > max_chars:
                out.append("... (truncated)")
                return "\n".join(out)[:max_chars]
    return "\n".join(out)[:max_chars]


def extract_analytics_data(paths: list[str], gemini_api_key: str,
                            model_name: str = "gemini-2.5-flash",
                            log=None) -> str:
    """
    Extract usable analytics from uploaded files.
      • Excel → openpyxl text dump
      • PDF / PNG / JPG → Gemini multimodal vision/PDF reading

    Returns a single text block with bullet-point analytics summaries.
    """
    if not paths:
        return ""

    summaries = []

    # ─── Excel (no API needed) ───
    excel_paths = [p for p in paths if Path(p).suffix.lower() in (".xlsx", ".xlsm")]
    for p in excel_paths:
        try:
            txt = _extract_excel_text(Path(p))
            if txt.strip():
                summaries.append(f"=== {Path(p).name} (Excel) ===\n{txt[:8000]}")
                if log:
                    log(f"  ✓ Read Excel: {Path(p).name}")
        except Exception as e:
            if log:
                log(f"  ⚠ Excel read failed for {Path(p).name}: {e}")

    # ─── PDF / images via Gemini multimodal ───
    visual_paths = [p for p in paths
                     if Path(p).suffix.lower() in (".pdf", ".png", ".jpg", ".jpeg", ".webp")]
    if visual_paths and gemini_api_key:
        try:
            import google.genai as _genai
            from google.genai import types as _genai_types
            _gclient = _genai.Client(api_key=gemini_api_key)
        except Exception as e:
            if log:
                log(f"  ⚠ Gemini init failed: {e}")
            return "\n\n".join(summaries)

        for p in visual_paths:
            path = Path(p)
            if not path.exists():
                continue
            try:
                if log:
                    log(f"  ⟳ Reading {path.name} via Gemini Vision…")
                up = _gclient.files.upload(file=str(path), config={"display_name": path.name})
                # Wait for ACTIVE
                for _ in range(30):
                    if up.state.name == "ACTIVE":
                        break
                    if up.state.name == "FAILED":
                        raise RuntimeError("upload failed")
                    time.sleep(1)
                    up = _gclient.files.get(name=up.name)

                config = _genai_types.GenerateContentConfig(
                    temperature=0.2,
                    response_mime_type="application/json",
                )
                resp = _gclient.models.generate_content(
                    model=model_name,
                    contents=[
                        ANALYTICS_SUMMARY_PROMPT,
                        "\n\nFile attached. Extract all concrete data points.",
                        up,
                    ],
                    config=config,
                )
                text = resp.text.strip()
                text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text,
                                flags=re.MULTILINE).strip()
                data = json.loads(text)
                summary = data.get("analytics_summary", "")
                if summary:
                    summaries.append(f"=== {path.name} (extracted by AI) ===\n{summary[:6000]}")
                    if log:
                        log(f"     ✓ Got {len(summary)} chars of summary")
            except Exception as e:
                if log:
                    log(f"  ⚠ Failed to read {path.name}: {e}")
                continue

    return "\n\n".join(summaries)


# ─────────────────────────────────────────────────────────────────────────
# Reference materials reader (PDFs / images via brand_extractor logic)
# ─────────────────────────────────────────────────────────────────────────
def extract_references_text(reference_paths: list[str], gemini_key: str,
                             model_name: str = "gemini-2.5-flash", log=None) -> str:
    """Read uploaded reference files (PDF/PPTX/DOCX/TXT) and return concatenated text."""
    if not reference_paths:
        return ""
    try:
        import brand_extractor as bx
    except ImportError:
        return ""
    parts = []
    for fp in reference_paths:
        p = Path(fp)
        if not p.exists():
            continue
        ext = p.suffix.lower()
        if ext == ".pdf":
            # Use Gemini for PDF text extraction via the brand_extractor pipeline
            # but simpler: just pass through to LLM later. For now, skip PDF heavy reading.
            parts.append(f"=== {p.name} (PDF referenced — content via filename only) ===")
        elif ext in (".pptx", ".ppt"):
            parts.append(f"=== {p.name} ===\n{bx._extract_pptx_text(p)}")
        elif ext in (".docx", ".doc"):
            parts.append(f"=== {p.name} ===\n{bx._extract_docx_text(p)}")
        elif ext in (".txt", ".md"):
            parts.append(f"=== {p.name} ===\n{bx._extract_text_file(p)}")
    return "\n\n".join(parts)[:8000]


# ─────────────────────────────────────────────────────────────────────────
# Storage helpers
# ─────────────────────────────────────────────────────────────────────────
def strategy_json_path(brands_dir: Path, brand_key: str) -> Path:
    """Path to the LATEST plan (always overwritten with newest version)."""
    return brands_dir / f"{brand_key}_strategy.json"


def versioned_strategy_path(brands_dir: Path, brand_key: str, version: int) -> Path:
    return brands_dir / f"{brand_key}_strategy_v{version}.json"


def list_plan_versions(brands_dir: Path, brand_key: str) -> list[int]:
    """Return sorted list of existing version numbers for this brand."""
    versions = []
    for p in brands_dir.glob(f"{brand_key}_strategy_v*.json"):
        m = re.search(r"_strategy_v(\d+)\.json$", p.name)
        if m:
            try:
                versions.append(int(m.group(1)))
            except Exception:
                pass
    return sorted(versions)


def next_plan_version(brands_dir: Path, brand_key: str) -> int:
    """v1 if no history, else previous+1."""
    versions = list_plan_versions(brands_dir, brand_key)
    return (versions[-1] + 1) if versions else 1


def save_strategy(brands_dir: Path, brand_key: str, strategy: dict) -> Path:
    """Save both the latest pointer + a versioned snapshot."""
    version = (strategy.get("_meta") or {}).get("version") or 1
    # Latest pointer
    latest = strategy_json_path(brands_dir, brand_key)
    latest.write_text(json.dumps(strategy, indent=2, ensure_ascii=False), encoding="utf-8")
    # Versioned snapshot
    snap = versioned_strategy_path(brands_dir, brand_key, version)
    snap.write_text(json.dumps(strategy, indent=2, ensure_ascii=False), encoding="utf-8")
    return latest


def load_strategy(brands_dir: Path, brand_key: str, version: int = None) -> dict | None:
    """Load the latest plan, or a specific version if asked."""
    if version:
        p = versioned_strategy_path(brands_dir, brand_key, version)
    else:
        p = strategy_json_path(brands_dir, brand_key)
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None


def get_previous_plan_brief(brands_dir: Path, brand_key: str, current_version: int) -> str:
    """Short text brief of the previous plan version — used during plan revision."""
    if current_version <= 1:
        return ""
    prev = load_strategy(brands_dir, brand_key, version=current_version - 1)
    if not prev:
        return ""
    es = prev.get("executive_summary", {})
    pillars = prev.get("growth_pillars", []) or []
    cs = prev.get("content_strategy", {})
    out = [
        f"PREVIOUS PLAN — v{current_version - 1} (saved "
        + (prev.get("_meta", {}).get("generated_at") or "?")[:10] + ")",
        "",
        "Targets set in v" + str(current_version - 1) + ":",
        f"  • North star: {es.get('north_star_metric', '?')}",
        f"  • Headline target: {es.get('headline_target', '?')}",
        "",
        "Pillars chosen in v" + str(current_version - 1) + ":",
    ]
    for p in pillars[:6]:
        out.append(f"  • {p.get('name', '?')} ({p.get('share_of_calendar', '?')}%) — {p.get('thesis', '')[:120]}")
    if cs.get("format_mix"):
        out.append("")
        out.append("Format mix in v" + str(current_version - 1) + ": " +
                    ", ".join(f"{k} {v}%" for k, v in cs["format_mix"].items()))
    return "\n".join(out)[:3500]


def get_strategy_brief(brands_dir: Path, brand_key: str) -> str:
    """Compact strategy summary for Strategist agent prompt injection."""
    s = load_strategy(brands_dir, brand_key)
    if not s:
        return ""
    out = []
    es = s.get("executive_summary", {})
    if es.get("opportunity"):
        out.append("STRATEGIC OPPORTUNITY:\n  " + es["opportunity"])
    if es.get("headline_target"):
        out.append("\n90-DAY TARGET:\n  " + es["headline_target"])

    pillars = s.get("growth_pillars", []) or []
    if pillars:
        out.append("\nGROWTH PILLARS (allocate posts proportionally to share_of_calendar):")
        for p in pillars:
            out.append(f"  • {p.get('name','')} ({p.get('share_of_calendar','?')}%) — {p.get('thesis','')[:200]}")

    cs = s.get("content_strategy", {})
    if cs.get("format_mix"):
        mix = ", ".join(f"{k} {v}%" for k, v in cs["format_mix"].items())
        out.append(f"\nFORMAT MIX TARGET: {mix}")
    if cs.get("hook_strategy"):
        out.append(f"HOOK STRATEGY: {cs['hook_strategy']}")

    tw = s.get("trend_watch", []) or []
    if tw:
        out.append("\nTRENDS TO RIDE:")
        for t in tw[:5]:
            out.append(f"  • {t.get('trend','')} — {t.get('how_to_ride','')[:120]}")

    return "\n".join(out)[:4500]


# ─────────────────────────────────────────────────────────────────────────
# MAIN ENTRY — orchestrator
# ─────────────────────────────────────────────────────────────────────────
def detect_brand_mode(brand: dict, brands_dir: Path, brand_key: str) -> str:
    """'existing' if Meta is connected AND there's at least one Analyst report; else 'new'."""
    import meta_client as _mc
    has_meta = _mc.has_meta_credentials(brand)
    if not has_meta:
        return "new"
    try:
        import analyst_agent as _aa
        r = _aa.latest_report(brands_dir, brand_key)
        if r:
            return "existing"
    except Exception:
        pass
    return "new"


def run_growth_planner(
    brand_key: str,
    website_url: str,
    ongoing_campaigns: str,
    reference_paths: list[str],
    output_dir: str,
    llm_client,
    settings: dict,
    instagram_handle: str = "",
    facebook_handle: str = "",
    manual_analytics: str = "",
    analytics_paths: list[str] = None,
    force_mode: str = None,           # "existing" / "new" — override auto-detect
    progress_callback=None,
    log_callback=None,
) -> dict:
    """
    End-to-end: scrape website + Tavily + trends + LLM + PPT + JSON.
    Returns {"json_path": ..., "pptx_path": ..., "strategy": {...}}.
    """
    import strategy_ppt
    import competitor_analyzer

    brand = core.load_brand(brand_key)
    brands_dir = core.BRANDS_DIR

    # ─── Detect mode + version ───
    mode = force_mode or detect_brand_mode(brand, brands_dir, brand_key)
    plan_version = next_plan_version(brands_dir, brand_key)
    if log_callback:
        log_callback(f"\n▸ Brand mode: {mode.upper()}  ·  Generating plan v{plan_version}")

    steps = ["Detecting brand mode",
             "Scraping brand website",
             "Analysing own social media",
             "Loading Analyst Baba report" if mode == "existing" else "Skipping audit (new brand)",
             "Web research (Tavily)",
             "Harvesting trends",
             "Reading references",
             "Loading competitor playbook",
             "LLM strategist synthesis",
             "Saving outputs",
             "Building PPT deck"]

    def _step(idx, label):
        if progress_callback:
            progress_callback(idx, len(steps), label)
        if log_callback:
            log_callback(f"\n[{idx}/{len(steps)}] {label}")

    # Auto-load saved handles from brand JSON if not passed in
    saved_handles = brand.get("social_handles", {}) or {}
    if not instagram_handle and saved_handles.get("instagram"):
        instagram_handle = saved_handles["instagram"]
    if not facebook_handle and saved_handles.get("facebook"):
        facebook_handle = saved_handles["facebook"]

    # 1. Detecting brand mode (already detected above — log it for the user)
    _step(1, steps[0])
    if log_callback:
        log_callback(f"  ✓ Mode: {mode.upper()}  ·  Version: v{plan_version}")

    # 2. Scrape website
    _step(2, steps[1])
    website_data = scrape_brand_website(website_url, log=log_callback) if website_url else {"homepage": "", "pages": {}}

    # 3. Analyse own social
    _step(3, steps[2])
    self_social = analyse_self_social(
        brand_key=brand_key, brand=brand,
        ig_handle=instagram_handle, fb_handle=facebook_handle,
        settings=settings, llm_client=llm_client, log=log_callback)
    self_brief = format_self_social_brief(self_social)

    # 3b. Extract analytics files (Excel / PDF / PNG) — augments manual_analytics text
    if analytics_paths:
        if log_callback:
            log_callback(f"  ⟳ Reading {len(analytics_paths)} analytics file(s)…")
        extracted = extract_analytics_data(
            analytics_paths, settings.get("gemini_api_key", ""),
            settings.get("model", "gemini-2.5-flash"), log=log_callback)
        if extracted:
            if manual_analytics:
                manual_analytics = (
                    "USER NOTES:\n" + manual_analytics +
                    "\n\nEXTRACTED FROM UPLOADED FILES:\n" + extracted
                )
            else:
                manual_analytics = extracted
    # Persist handles to brand JSON for future runs
    if instagram_handle or facebook_handle:
        brand["social_handles"] = {
            "instagram": instagram_handle or saved_handles.get("instagram", ""),
            "facebook": facebook_handle or saved_handles.get("facebook", ""),
        }
        (brands_dir / f"{brand_key}.json").write_text(
            json.dumps(brand, indent=2, ensure_ascii=False), encoding="utf-8")

    # 4. Load Analyst Baba report + previous plan (existing mode only)
    _step(4, steps[3])
    analyst_brief = ""
    previous_plan_brief = ""
    if mode == "existing":
        try:
            import analyst_agent as _aa
            analyst_brief = _aa.get_analyst_brief_for_grook(brands_dir, brand_key)
            if log_callback:
                if analyst_brief:
                    log_callback(f"  ✓ Analyst report loaded ({len(analyst_brief)} chars)")
                else:
                    log_callback("  ⚠ No Analyst report yet — run Analyst Baba first for sharper plan")
        except Exception as e:
            if log_callback:
                log_callback(f"  ⚠ Analyst load failed (non-fatal): {e}")
        if plan_version > 1:
            previous_plan_brief = get_previous_plan_brief(brands_dir, brand_key, plan_version)
            if log_callback and previous_plan_brief:
                log_callback(f"  ✓ Previous plan v{plan_version - 1} loaded for delta comparison")
    else:
        if log_callback:
            log_callback("  · New brand — no audit data exists yet")

    # 5. Tavily
    _step(5, steps[4])
    tavily_data = tavily_research(brand, settings.get("tavily_api_key", ""), log=log_callback)

    # 6. Trends
    _step(6, steps[5])
    try:
        trend_scout.refresh_trend_snapshot(
            brands_dir=brands_dir, brand_key=brand_key, brand=brand,
            settings=settings, llm_client=llm_client, log=log_callback)
    except Exception as e:
        if log_callback:
            log_callback(f"  ⚠ Trend refresh issue (non-fatal): {e}")
    trends_brief = trend_scout.get_trend_brief(brands_dir, brand_key)

    # 7. References (incl. brand assets folder)
    _step(7, steps[6])
    extended_refs = list(reference_paths or [])
    assets = brand.get("assets", {}) or {}
    for p in (assets.get("guidelines", []) or []) + (assets.get("references", []) or []):
        full = brands_dir / p
        if full.exists():
            extended_refs.append(str(full))
    refs_text = extract_references_text(extended_refs,
                                          settings.get("gemini_api_key", ""),
                                          settings.get("model", "gemini-2.5-flash"),
                                          log=log_callback)

    # 8. Competitors
    _step(8, steps[7])
    competitor_playbook = competitor_analyzer.get_playbook_text(brands_dir, brand_key)
    if log_callback and competitor_playbook:
        log_callback("  ✓ Competitor playbook loaded.")

    # 9. LLM synthesis (the slow step — progress bar will rest here for ~30-60s)
    _step(9, steps[8])
    if log_callback:
        log_callback("  ⏳ This step takes 30-90 seconds while the LLM reasons.")
    strategy = generate_strategy_json(
        brand=brand, website_data=website_data,
        ongoing_campaigns_text=ongoing_campaigns or "",
        references_text=refs_text,
        tavily_data=tavily_data,
        trends_brief=trends_brief,
        competitor_playbook=competitor_playbook,
        self_social_brief=self_brief,
        manual_analytics=manual_analytics or "",
        analyst_brief=analyst_brief,
        previous_plan_brief=previous_plan_brief,
        brand_mode=mode,
        plan_version=plan_version,
        llm_client=llm_client, log=log_callback)
    # Embed the current state snapshot
    strategy["current_social_state"] = self_social
    # Attach metadata (with versioning + mode)
    strategy["_meta"] = {
        "brand_key": brand_key,
        "brand_mode": mode,
        "version": plan_version,
        "previous_version": plan_version - 1 if plan_version > 1 else None,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "website_url": website_url,
        "scraped_pages": list(website_data.get("pages", {}).keys()),
        "tavily_queries": tavily_data.get("queries", []),
        "analyst_report_used": bool(analyst_brief),
    }

    # 10. Save JSON FIRST so it's never lost even if PPT fails
    _step(10, steps[9])
    json_path = save_strategy(brands_dir, brand_key, strategy)
    if log_callback:
        log_callback(f"  ✓ Strategy v{plan_version} saved: {json_path.name}")

    # 11. Build PPT
    _step(11, steps[10])
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = "".join(c for c in brand.get("name", brand_key) if c.isalnum() or c in "_-").strip("_")
    pptx_path = out_dir / f"{safe}_Strategy_{ts}.pptx"

    # Use brand's primary logo if present
    brand_logo = None
    assets = brand.get("assets", {}) or {}
    if assets.get("logo_primary"):
        cand = brands_dir / assets["logo_primary"]
        if cand.exists():
            brand_logo = str(cand)

    try:
        strategy_ppt.build_strategy_pptx(strategy, brand, settings, str(pptx_path),
                                           brand_logo=brand_logo)
        if log_callback:
            log_callback(f"  ✓ PPT saved: {pptx_path.name}")
    except Exception as e:
        if log_callback:
            log_callback(f"  ⚠ PPT build failed (JSON still saved): {e}")
        raise

    return {"json_path": str(json_path), "pptx_path": str(pptx_path), "strategy": strategy}

